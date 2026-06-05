#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
v11.2 Deep Holding Analysis
- Reads manual holdings/trade logs from repository root.
- Builds richer holding decision tables for HTML, Excel, and Google Sheets CSV.
- Designed to be safe: it never modifies manual input files.
"""

from __future__ import annotations

import csv
import html
import math
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    load_workbook = None

ROOT = Path('.')
DATA_DIR = ROOT / 'docs' / 'data'
PAGE_DIR = ROOT / 'docs' / 'v11_holdings'
TODAY = datetime.now().strftime('%Y-%m-%d')
TODAY_KEY = datetime.now().strftime('%Y%m%d')

HOLDING_FILES = [
    'holdings_manual_input.csv',
    '보유종목_수동입력.csv',
]
TRADE_FILES = [
    'trade_log_manual_input.csv',
    '매매기록_수동입력.csv',
]

COLUMN_ALIASES = {
    'status': ['status', '상태'],
    'stock_name': ['stock_name', '종목명', 'name'],
    'stock_code': ['stock_code', '종목코드', 'code'],
    'quantity': ['quantity', '보유수량', '수량', 'qty'],
    'avg_price': ['avg_price', '평균단가', '매입단가', '평단'],
    'buy_date': ['buy_date', '매수일', '진입일'],
    'strategy': ['strategy', '전략구분'],
    'target_price': ['target_price', '목표가'],
    'stop_loss': ['stop_loss', '손절가'],
    'weight_note': ['weight_note', '비중메모'],
    'memo': ['memo', '메모'],
    'current_price': ['current_price', '현재가', '평가단가'],
}

TRADE_ALIASES = {
    'trade_date': ['trade_date', '거래일', '날짜'],
    'trade_type': ['trade_type', '구분', '매매구분'],
    'stock_name': ['stock_name', '종목명', 'name'],
    'stock_code': ['stock_code', '종목코드', 'code'],
    'quantity': ['quantity', '수량', 'qty'],
    'price': ['price', '단가', '체결가'],
    'fee': ['fee', '수수료'],
    'tax': ['tax', '세금'],
    'memo': ['memo', '메모'],
}


def log(msg: str) -> None:
    print(f'[v11.2] {msg}')


def find_first_existing(names: List[str]) -> Optional[Path]:
    for name in names:
        p = ROOT / name
        if p.exists():
            return p
    return None


def read_csv_smart(path: Optional[Path]) -> pd.DataFrame:
    if path is None or not path.exists():
        return pd.DataFrame()
    for enc in ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr']:
        try:
            return pd.read_csv(path, encoding=enc, dtype=str).fillna('')
        except Exception:
            continue
    log(f'CSV 읽기 실패: {path}')
    return pd.DataFrame()


def normalize_columns(df: pd.DataFrame, aliases: Dict[str, List[str]]) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    stripped = {str(c).strip(): c for c in df.columns}
    rename = {}
    for canonical, candidates in aliases.items():
        for cand in candidates:
            if cand in stripped:
                rename[stripped[cand]] = canonical
                break
    df = df.rename(columns=rename)
    for canonical in aliases:
        if canonical not in df.columns:
            df[canonical] = ''
    return df


def clean_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in {'nan', 'none', '-'}:
        return None
    s = s.replace(',', '').replace('원', '').replace('%', '').replace('+', '').strip()
    s = re.sub(r'[^0-9.\-]', '', s)
    if s in {'', '-', '.', '-.'}:
        return None
    try:
        return float(s)
    except Exception:
        return None


def clean_int(v: Any) -> int:
    x = clean_number(v)
    return int(x) if x is not None and not math.isnan(x) else 0


def fmt_price(v: Any) -> str:
    x = clean_number(v)
    if x is None:
        return '-'
    return f'{int(round(x)):,}원'


def fmt_pct(v: Any) -> str:
    x = clean_number(v)
    if x is None:
        return '-'
    return f'{x:+.2f}%'


def make_key(name: Any, code: Any = '') -> str:
    code_s = str(code or '').strip().zfill(6) if str(code or '').strip().isdigit() else str(code or '').strip()
    name_s = str(name or '').strip()
    return code_s or name_s


def parse_date(v: Any) -> Optional[date]:
    s = str(v or '').strip()
    if not s:
        return None
    for fmt in ['%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d', '%Y%m%d']:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def find_latest_xlsx() -> Optional[Path]:
    files = [p for p in ROOT.glob('*.xlsx') if not p.name.startswith('~$')]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def read_sheet_if_exists(xlsx: Optional[Path], sheet_names: List[str]) -> pd.DataFrame:
    if xlsx is None or not xlsx.exists():
        return pd.DataFrame()
    try:
        xl = pd.ExcelFile(xlsx)
        for s in sheet_names:
            if s in xl.sheet_names:
                return pd.read_excel(xlsx, sheet_name=s, dtype=str).fillna('')
    except Exception as e:
        log(f'엑셀 시트 읽기 실패: {xlsx} / {e}')
    return pd.DataFrame()


def read_candidates(xlsx: Optional[Path]) -> pd.DataFrame:
    # Prefer Google Sheets CSV output if already generated.
    csv_path = DATA_DIR / 'latest_candidates.csv'
    if csv_path.exists():
        return read_csv_smart(csv_path)
    return read_sheet_if_exists(xlsx, ['TOP후보_요약', '추천 리스트', '추천리스트'])


def read_news(xlsx: Optional[Path]) -> pd.DataFrame:
    csv_path = DATA_DIR / 'latest_news_summary.csv'
    if csv_path.exists():
        return read_csv_smart(csv_path)
    return read_sheet_if_exists(xlsx, ['네이버뉴스_요약', '뉴스_요약'])


def value_from_row(row: pd.Series, candidates: List[str]) -> str:
    for c in candidates:
        if c in row.index and str(row.get(c, '')).strip():
            return str(row.get(c, '')).strip()
    return ''


def build_candidate_lookup(cand_df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    if cand_df.empty:
        return out
    for _, row in cand_df.iterrows():
        name = value_from_row(row, ['stock_name', '종목명', '종목', 'name'])
        code = value_from_row(row, ['stock_code', '종목코드', '코드', 'code'])
        key = make_key(name, code)
        if not key:
            continue
        score = value_from_row(row, ['추천점수', 'score', '점수', '종합점수'])
        rank = value_from_row(row, ['순위', 'rank'])
        source = value_from_row(row, ['후보출처', 'source', '추천출처'])
        price = value_from_row(row, ['현재가', 'current_price', '종가', '기준가'])
        out[key] = {
            'candidate_name': name,
            'candidate_code': code,
            'recommend_score': score,
            'recommend_rank': rank,
            'candidate_source': source,
            'candidate_price': price,
            'candidate_matched': 'Y',
        }
        if name:
            out[name] = out[key]
    return out


def count_news_for_stock(news_df: pd.DataFrame, stock_name: str) -> Tuple[int, str]:
    if news_df.empty or not stock_name:
        return 0, ''
    text_cols = [c for c in news_df.columns if any(k in str(c).lower() for k in ['title', '제목', 'summary', '요약', 'query', '검색', '키워드'])]
    if not text_cols:
        text_cols = list(news_df.columns)
    matched = []
    for _, row in news_df.iterrows():
        joined = ' '.join(str(row.get(c, '')) for c in text_cols)
        if stock_name in joined:
            title = value_from_row(row, ['title', '제목', 'news_title']) or joined[:80]
            matched.append(title)
    return len(matched), ' | '.join(matched[:2])


def trade_summary(trade_df: pd.DataFrame, key: str, name: str) -> Dict[str, Any]:
    if trade_df.empty:
        return {'last_trade_date': '', 'last_trade_type': '', 'buy_count': 0, 'sell_count': 0}
    matches = []
    for _, row in trade_df.iterrows():
        r_key = make_key(row.get('stock_name', ''), row.get('stock_code', ''))
        if r_key == key or str(row.get('stock_name', '')).strip() == name:
            matches.append(row)
    if not matches:
        return {'last_trade_date': '', 'last_trade_type': '', 'buy_count': 0, 'sell_count': 0}
    mdf = pd.DataFrame(matches)
    buy_count = int(mdf['trade_type'].astype(str).str.lower().str.contains('buy|매수').sum()) if 'trade_type' in mdf else 0
    sell_count = int(mdf['trade_type'].astype(str).str.lower().str.contains('sell|매도').sum()) if 'trade_type' in mdf else 0
    last = mdf.iloc[-1]
    return {
        'last_trade_date': str(last.get('trade_date', '')),
        'last_trade_type': str(last.get('trade_type', '')),
        'buy_count': buy_count,
        'sell_count': sell_count,
    }


def decide_action(pl_pct: Optional[float], current: Optional[float], target: Optional[float], stop: Optional[float], candidate_matched: bool, news_count: int) -> Tuple[str, str, str]:
    risk = 'NORMAL'
    if pl_pct is None:
        return 'NO_PRICE', '현재가/평균단가가 없어 수익률 판단은 보류합니다.', 'UNKNOWN'
    if stop is not None and current is not None and current <= stop:
        return 'STOP_WATCH', '손절 기준가에 도달했거나 이탈했습니다. 즉시 재점검이 필요합니다.', 'HIGH'
    if pl_pct <= -5:
        return 'STOP_WATCH', '손실률이 -5% 이하입니다. 손절 기준과 보유 이유를 다시 확인하세요.', 'HIGH'
    if target is not None and current is not None and current >= target:
        return 'TAKE_PROFIT_2', '목표가에 도달했습니다. 분할 익절 또는 추세 유지 여부를 확인하세요.', 'LOW'
    if pl_pct >= 8:
        return 'TAKE_PROFIT_2', '수익률이 +8% 이상입니다. 일부 익절 후 잔여 물량 추세 대응을 고려하세요.', 'LOW'
    if pl_pct >= 5:
        return 'TAKE_PROFIT_1', '1차 익절 구간입니다. 일부 익절 또는 손절가 상향을 검토하세요.', 'LOW'
    if pl_pct <= -3:
        return 'RISK_CHECK', '손실 구간입니다. 추가매수보다 손절/관망 기준을 먼저 확인하세요.', 'MEDIUM'
    if candidate_matched and news_count > 0:
        return 'HOLD_ADD_WATCH', '추천후보와 뉴스에 함께 잡힙니다. 장중 거래량 확인 후 추가매수 후보로 관찰하세요.', 'NORMAL'
    if candidate_matched:
        return 'HOLD_WATCH', '추천후보에 포함되어 있습니다. 유지 관점에서 가격/거래량을 확인하세요.', 'NORMAL'
    return 'HOLD', '특별한 매도 신호는 약하지만 신규 모멘텀도 강하지 않습니다. 보유/관망 우선입니다.', 'NORMAL'


def build_deep_analysis() -> Tuple[pd.DataFrame, pd.DataFrame, Optional[Path]]:
    holdings_path = find_first_existing(HOLDING_FILES)
    trade_path = find_first_existing(TRADE_FILES)
    holdings = normalize_columns(read_csv_smart(holdings_path), COLUMN_ALIASES)
    trades = normalize_columns(read_csv_smart(trade_path), TRADE_ALIASES)
    xlsx = find_latest_xlsx()
    candidates = read_candidates(xlsx)
    news = read_news(xlsx)
    cand_lookup = build_candidate_lookup(candidates)

    log(f'holdings file: {holdings_path or "NOT FOUND"}')
    log(f'trade log file: {trade_path or "NOT FOUND"}')
    log(f'latest xlsx: {xlsx or "NOT FOUND"}')
    log(f'holdings rows: {len(holdings)} / trade rows: {len(trades)} / candidate rows: {len(candidates)} / news rows: {len(news)}')

    if holdings.empty:
        deep = pd.DataFrame([{
            'analysis_date': TODAY,
            'stock_name': 'NO_HOLDINGS_FILE',
            'decision': 'CHECK_INPUT',
            'risk_level': 'UNKNOWN',
            'guide_summary': 'holdings_manual_input.csv 파일이 저장소 루트에 있는지 확인하세요.',
        }])
        return deep, pd.DataFrame(), xlsx

    rows = []
    guide_rows = []
    for _, row in holdings.iterrows():
        name = str(row.get('stock_name', '')).strip()
        code = str(row.get('stock_code', '')).strip()
        key = make_key(name, code)
        qty = clean_int(row.get('quantity', ''))
        avg = clean_number(row.get('avg_price', ''))
        target = clean_number(row.get('target_price', ''))
        stop = clean_number(row.get('stop_loss', ''))
        if target is None and avg is not None:
            target = avg * 1.08
        if stop is None and avg is not None:
            stop = avg * 0.95

        cand = cand_lookup.get(key) or cand_lookup.get(name) or {}
        current = clean_number(row.get('current_price', '')) or clean_number(cand.get('candidate_price', '')) or avg
        pl_pct = ((current - avg) / avg * 100) if current is not None and avg not in (None, 0) else None
        eval_amount = current * qty if current is not None else None
        cost_amount = avg * qty if avg is not None else None
        unrealized_pl = (eval_amount - cost_amount) if eval_amount is not None and cost_amount is not None else None
        n_count, n_titles = count_news_for_stock(news, name)
        t_sum = trade_summary(trades, key, name)
        buy_dt = parse_date(row.get('buy_date', ''))
        holding_days = (date.today() - buy_dt).days if buy_dt else ''
        candidate_matched = bool(cand.get('candidate_matched'))
        decision, guide, risk = decide_action(pl_pct, current, target, stop, candidate_matched, n_count)
        dist_target = ((target - current) / current * 100) if target is not None and current not in (None, 0) else None
        dist_stop = ((current - stop) / current * 100) if stop is not None and current not in (None, 0) else None

        entry_guide = '신규 매수보다는 보유 대응 우선'
        if decision in {'HOLD_ADD_WATCH', 'HOLD_WATCH'}:
            entry_guide = '거래량 증가와 지지 확인 시 추가매수 검토'
        elif decision.startswith('TAKE'):
            entry_guide = '신규/추가매수보다 익절 기준 관리 우선'
        elif decision in {'STOP_WATCH', 'RISK_CHECK'}:
            entry_guide = '추가매수 금지, 손절/관망 기준 우선'

        profit_guide = '목표가 또는 +5~8% 구간에서 분할 익절 검토'
        stop_guide = '손절가 또는 -3~-5% 구간 이탈 시 비중 축소 검토'
        if target is not None:
            profit_guide = f'1차 익절: {fmt_price(target)} 부근 / 강한 추세면 일부만 익절'
        if stop is not None:
            stop_guide = f'손절/경고: {fmt_price(stop)} 이탈 시 재검토'

        rows.append({
            'analysis_date': TODAY,
            'stock_name': name,
            'stock_code': code,
            'status': row.get('status', ''),
            'quantity': qty,
            'avg_price': avg,
            'current_price': current,
            'target_price': target,
            'stop_loss': stop,
            'unrealized_pl': unrealized_pl,
            'unrealized_pl_pct': pl_pct,
            'distance_to_target_pct': dist_target,
            'distance_to_stop_pct': dist_stop,
            'holding_days': holding_days,
            'decision': decision,
            'risk_level': risk,
            'guide_summary': guide,
            'entry_guide': entry_guide,
            'profit_guide': profit_guide,
            'stop_guide': stop_guide,
            'candidate_matched': 'Y' if candidate_matched else 'N',
            'recommend_rank': cand.get('recommend_rank', ''),
            'recommend_score': cand.get('recommend_score', ''),
            'candidate_source': cand.get('candidate_source', ''),
            'news_count': n_count,
            'news_titles': n_titles,
            'last_trade_date': t_sum['last_trade_date'],
            'last_trade_type': t_sum['last_trade_type'],
            'buy_count': t_sum['buy_count'],
            'sell_count': t_sum['sell_count'],
            'memo': row.get('memo', ''),
        })
        guide_rows.append({
            'stock_name': name,
            'stock_code': code,
            'decision': decision,
            'risk_level': risk,
            'entry_guide': entry_guide,
            'profit_guide': profit_guide,
            'stop_guide': stop_guide,
            'action_today': guide,
        })

    deep_df = pd.DataFrame(rows)
    guide_df = pd.DataFrame(guide_rows)
    return deep_df, guide_df, xlsx


def save_csv_outputs(deep_df: pd.DataFrame, guide_df: pd.DataFrame) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    deep_df.to_csv(DATA_DIR / 'latest_holding_deep_analysis.csv', index=False, encoding='utf-8-sig')
    guide_df.to_csv(DATA_DIR / 'latest_holding_action_guide.csv', index=False, encoding='utf-8-sig')
    log(f'CSV outputs saved: {DATA_DIR}')


def style_sheet(ws) -> None:
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value or '')) for c in col[:100])
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 38))
    ws.freeze_panes = 'A2'


def update_workbook(deep_df: pd.DataFrame, guide_df: pd.DataFrame, xlsx: Optional[Path]) -> None:
    if xlsx is None or load_workbook is None:
        log('엑셀 파일이 없거나 openpyxl 사용 불가: 엑셀 시트 업데이트 생략')
        return
    try:
        wb = load_workbook(xlsx)
        for sheet_name, df in [('보유종목_심화분석', deep_df), ('보유대응_가이드', guide_df)]:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            ws.append(list(df.columns))
            for _, r in df.iterrows():
                ws.append([r.get(c, '') for c in df.columns])
            style_sheet(ws)
        wb.save(xlsx)
        log(f'엑셀 업데이트 완료: {xlsx}')
    except Exception as e:
        log(f'엑셀 업데이트 실패: {e}')


def badge(decision: str) -> str:
    cls = 'neutral'
    if decision.startswith('TAKE'):
        cls = 'profit'
    elif decision in {'STOP_WATCH', 'RISK_CHECK'}:
        cls = 'risk'
    elif decision in {'HOLD_ADD_WATCH', 'HOLD_WATCH'}:
        cls = 'watch'
    elif decision == 'HOLD':
        cls = 'hold'
    return f'<span class="badge {cls}">{html.escape(decision)}</span>'


def render_html(deep_df: pd.DataFrame, guide_df: pd.DataFrame) -> None:
    PAGE_DIR.mkdir(parents=True, exist_ok=True)
    if deep_df.empty:
        cards = '<p>보유종목 분석 데이터가 없습니다.</p>'
    else:
        cards_list = []
        for _, r in deep_df.iterrows():
            name = html.escape(str(r.get('stock_name', '')))
            code = html.escape(str(r.get('stock_code', '')))
            decision = str(r.get('decision', ''))
            pl = fmt_pct(r.get('unrealized_pl_pct'))
            news_count = html.escape(str(r.get('news_count', '')))
            risk = html.escape(str(r.get('risk_level', '')))
            cards_list.append(f'''
<section class="card">
  <div class="card-head">
    <div><h2>{name}</h2><p class="sub">{code} · 보유수량 {html.escape(str(r.get('quantity','')))}주</p></div>
    <div>{badge(decision)}</div>
  </div>
  <div class="kpi-grid">
    <div><span>평균단가</span><b>{fmt_price(r.get('avg_price'))}</b></div>
    <div><span>기준가</span><b>{fmt_price(r.get('current_price'))}</b></div>
    <div><span>손익률</span><b>{pl}</b></div>
    <div><span>리스크</span><b>{risk}</b></div>
  </div>
  <table class="guide"><tbody>
    <tr><th>오늘 판단</th><td>{html.escape(str(r.get('guide_summary','')))}</td></tr>
    <tr><th>진입/추가매수</th><td>{html.escape(str(r.get('entry_guide','')))}</td></tr>
    <tr><th>익절 가이드</th><td>{html.escape(str(r.get('profit_guide','')))}</td></tr>
    <tr><th>손절 가이드</th><td>{html.escape(str(r.get('stop_guide','')))}</td></tr>
    <tr><th>추천후보 매칭</th><td>{html.escape(str(r.get('candidate_matched','')))} / 순위 {html.escape(str(r.get('recommend_rank','')))} / 점수 {html.escape(str(r.get('recommend_score','')))}</td></tr>
    <tr><th>뉴스 매칭</th><td>{news_count}건 · {html.escape(str(r.get('news_titles',''))[:160])}</td></tr>
    <tr><th>메모</th><td>{html.escape(str(r.get('memo','')))}</td></tr>
  </tbody></table>
</section>
''')
        cards = '\n'.join(cards_list)

    html_doc = f'''<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>v11.2 보유종목 심화분석</title>
<style>
  body {{ margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; background:#f5f7fb; color:#152033; }}
  header {{ padding:22px 18px; background:linear-gradient(135deg,#13294b,#1f6f8b); color:white; }}
  header h1 {{ margin:0 0 8px; font-size:23px; }}
  header p {{ margin:0; opacity:.88; font-size:14px; }}
  main {{ padding:16px; max-width:980px; margin:0 auto; }}
  .links {{ display:flex; gap:8px; flex-wrap:wrap; margin:14px 0 18px; }}
  .links a {{ text-decoration:none; background:white; color:#1f4e78; padding:9px 12px; border-radius:999px; border:1px solid #d8e3ef; font-weight:700; font-size:13px; }}
  .card {{ background:white; border:1px solid #e2e8f0; border-radius:18px; padding:16px; margin:14px 0; box-shadow:0 6px 18px rgba(15,23,42,.06); }}
  .card-head {{ display:flex; justify-content:space-between; align-items:flex-start; gap:12px; margin-bottom:12px; }}
  h2 {{ margin:0; font-size:20px; }}
  .sub {{ margin:4px 0 0; color:#64748b; font-size:13px; }}
  .badge {{ display:inline-block; padding:7px 10px; border-radius:999px; font-size:12px; font-weight:800; white-space:nowrap; }}
  .profit {{ background:#dbeafe; color:#1d4ed8; }}
  .risk {{ background:#fee2e2; color:#b91c1c; }}
  .watch {{ background:#fef3c7; color:#92400e; }}
  .hold {{ background:#dcfce7; color:#166534; }}
  .neutral {{ background:#e2e8f0; color:#334155; }}
  .kpi-grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:8px; margin:10px 0 14px; }}
  .kpi-grid div {{ background:#f8fafc; border:1px solid #e2e8f0; border-radius:12px; padding:10px; }}
  .kpi-grid span {{ display:block; color:#64748b; font-size:12px; margin-bottom:4px; }}
  .kpi-grid b {{ font-size:15px; }}
  table.guide {{ width:100%; border-collapse:collapse; overflow:hidden; border-radius:12px; font-size:14px; }}
  .guide th {{ width:120px; text-align:left; background:#f1f5f9; color:#334155; padding:10px; border:1px solid #e2e8f0; }}
  .guide td {{ padding:10px; border:1px solid #e2e8f0; line-height:1.45; }}
  .note {{ background:#fff7ed; border:1px solid #fed7aa; border-radius:14px; padding:12px; color:#7c2d12; font-size:13px; line-height:1.5; }}
  @media (max-width:720px) {{ .kpi-grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }} .card {{ padding:14px; }} .guide th {{ width:92px; }} }}
</style>
</head>
<body>
<header>
  <h1>v11.2 보유종목 심화분석</h1>
  <p>생성일: {html.escape(TODAY)} · 보유종목별 유지/익절/손절/추가매수 관찰 기준</p>
</header>
<main>
  <div class="links">
    <a href="../mobile/">모바일 홈</a>
    <a href="../latest/">최신 리포트</a>
    <a href="../v11_dashboard/">v11 대시보드</a>
    <a href="../data/latest_holding_deep_analysis.csv">CSV 보기</a>
  </div>
  <div class="note">이 화면은 투자 판단 보조용입니다. 리포트의 판단은 매수/매도 지시가 아니라, 장중 거래량·가격 흐름·뉴스 지속성 확인을 돕기 위한 기준표입니다.</div>
  {cards}
</main>
</body>
</html>'''
    (PAGE_DIR / 'index.html').write_text(html_doc, encoding='utf-8')
    log(f'HTML page saved: {PAGE_DIR / "index.html"}')


def append_mobile_links() -> None:
    # Keep this deliberately conservative: add a tiny link block only if files exist and marker is absent.
    for path in [ROOT / 'docs' / 'mobile' / 'index.html', ROOT / 'docs' / 'latest' / 'index.html']:
        if not path.exists():
            continue
        try:
            text = path.read_text(encoding='utf-8')
            if 'v11.2 보유종목 심화분석' in text:
                continue
            block = '''\n<div style="margin:16px 0;padding:14px;border:1px solid #dbeafe;border-radius:14px;background:#eff6ff;">
  <b>v11.2 보유종목 심화분석</b><br>
  <a href="../v11_holdings/" style="color:#1d4ed8;font-weight:700;">보유종목 익절·손절·추가매수 가이드 보기</a>
</div>\n'''
            if '</main>' in text:
                text = text.replace('</main>', block + '</main>')
            elif '</body>' in text:
                text = text.replace('</body>', block + '</body>')
            else:
                text += block
            path.write_text(text, encoding='utf-8')
            log(f'link injected: {path}')
        except Exception as e:
            log(f'link injection skipped: {path} / {e}')


def main() -> int:
    deep_df, guide_df, xlsx = build_deep_analysis()
    save_csv_outputs(deep_df, guide_df)
    update_workbook(deep_df, guide_df, xlsx)
    render_html(deep_df, guide_df)
    append_mobile_links()
    log('v11.2 deep holding analysis completed')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
