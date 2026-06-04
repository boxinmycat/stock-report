#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
v11.0 추천성과 검증 + 보유종목 판단 대시보드

- 추천성과 추적 CSV 누적
- 보유종목 HOLD/익절/손절 판단
- 진입/익절/손절 가이드 생성
- 엑셀 시트 3개 추가
- HTML 모바일 섹션 삽입
"""
from __future__ import annotations

import csv
import html
import math
import os
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KST = timezone(timedelta(hours=9))
ROOT = Path('.')
DOCS = ROOT / 'docs'
DATA = DOCS / 'data'
V11_PAGE = DOCS / 'v11_dashboard'
SITE_BASE_URL = os.getenv('SITE_BASE_URL', 'https://boxinmycat.github.io/stock-report').rstrip('/')

TRACKING_CSV = DATA / 'recommendation_tracking.csv'
LATEST_PERF_CSV = DATA / 'latest_performance.csv'
LATEST_HOLDING_JUDGMENT_CSV = DATA / 'latest_holding_judgment.csv'
LATEST_ENTRY_GUIDE_CSV = DATA / 'latest_entry_exit_guide.csv'

HOLDINGS_FILES = [
    os.getenv('HOLDINGS_MANUAL_CSV_FILE', ''),
    os.getenv('HOLDINGS_CSV_FILE', ''),
    'holdings_manual_input.csv',
    '보유종목_수동입력.csv',
]


def now_kst() -> datetime:
    return datetime.now(KST)


def session_label() -> str:
    env = os.getenv('REPORT_SESSION', '').upper().strip()
    if env in {'AM', 'PM', 'MANUAL'}:
        return env
    return 'AM' if now_kst().hour < 12 else 'PM'


def clean_text(v: Any) -> str:
    if v is None:
        return ''
    try:
        if pd.isna(v):
            return ''
    except Exception:
        pass
    s = str(v).strip()
    if s.lower() in {'nan', 'none', 'nat'}:
        return ''
    s = re.sub(r'\n?Name:\s*\d+.*?dtype:\s*object', '', s, flags=re.S)
    return s.strip()


def to_number(v: Any) -> Optional[float]:
    s = clean_text(v)
    if not s:
        return None
    s = s.replace(',', '').replace('원', '').replace('%', '').replace('주', '').strip()
    s = re.sub(r'[^0-9.\-]', '', s)
    if not s or s in {'-', '.', '-.'}:
        return None
    try:
        x = float(s)
        if math.isnan(x):
            return None
        return x
    except Exception:
        return None


def fmt_int(v: Any) -> str:
    x = to_number(v)
    if x is None:
        return ''
    return f'{int(round(x)):,}'


def fmt_pct(v: Any) -> str:
    x = to_number(v)
    if x is None:
        return ''
    return f'{x:.2f}%'


def normalize_code(v: Any) -> str:
    s = clean_text(v)
    if not s:
        return ''
    s = s.replace('.0', '') if re.fullmatch(r'\d+\.0', s) else s
    if re.fullmatch(r'\d{1,6}', s):
        return s.zfill(6)
    m = re.search(r'\d{6}', s)
    if m:
        return m.group(0)
    return s.strip()


def key_of(name: Any, code: Any) -> str:
    code = normalize_code(code)
    if code:
        return code
    return clean_text(name).replace(' ', '').upper()


def read_csv_flexible(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    for enc in ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr']:
        try:
            df = pd.read_csv(path, encoding=enc, dtype=str)
            df.columns = [clean_text(c) for c in df.columns]
            return df.dropna(how='all')
        except Exception:
            continue
    print(f'⚠️ CSV read failed: {path}')
    return pd.DataFrame()


def read_first_existing_csv(candidates: List[str]) -> pd.DataFrame:
    seen = []
    for c in candidates:
        if not c or c in seen:
            continue
        seen.append(c)
        p = Path(c)
        if p.exists():
            print(f'📄 CSV loaded: {p}')
            return read_csv_flexible(p)
    return pd.DataFrame()


def find_latest_xlsx() -> Optional[Path]:
    patterns = ['20*.xlsx', 'docs/reports/**/*.xlsx', 'stock_report/**/*.xlsx']
    files: List[Path] = []
    for pat in patterns:
        files.extend([p for p in ROOT.glob(pat) if p.is_file() and not p.name.startswith('~$')])
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def read_sheet(xlsx: Optional[Path], sheet_names: List[str]) -> pd.DataFrame:
    if xlsx is None or not xlsx.exists():
        return pd.DataFrame()
    try:
        xl = pd.ExcelFile(xlsx, engine='openpyxl')
    except Exception as e:
        print(f'⚠️ Excel open failed: {xlsx} / {e}')
        return pd.DataFrame()
    target = None
    for name in sheet_names:
        if name in xl.sheet_names:
            target = name
            break
    if not target:
        return pd.DataFrame()
    try:
        df = pd.read_excel(xlsx, sheet_name=target, engine='openpyxl', dtype=object)
        df.columns = [clean_text(c) for c in df.columns]
        return df.dropna(how='all')
    except Exception as e:
        print(f'⚠️ Sheet read failed: {target} / {e}')
        return pd.DataFrame()


def find_col(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
    if df is None or df.empty:
        return None
    norm_aliases = [re.sub(r'[\s_\-/]', '', a).lower() for a in aliases]
    for col in df.columns:
        c = re.sub(r'[\s_\-/]', '', clean_text(col)).lower()
        for a in norm_aliases:
            if a and (a == c or a in c):
                return col
    return None


def row_pick(row: Dict[str, Any], names: List[str]) -> str:
    norm = {re.sub(r'[\s_\-/]', '', clean_text(k)).lower(): v for k, v in row.items()}
    for n in names:
        key = re.sub(r'[\s_\-/]', '', n).lower()
        if key in norm:
            return clean_text(norm[key])
    for n in names:
        key = re.sub(r'[\s_\-/]', '', n).lower()
        for k, v in norm.items():
            if key and key in k:
                return clean_text(v)
    return ''


def candidate_rows(xlsx: Optional[Path], max_rows: int = 20) -> List[Dict[str, Any]]:
    df = read_sheet(xlsx, ['TOP후보_요약', '추천 리스트', '추천리스트'])
    if df.empty:
        return []
    name_col = find_col(df, ['종목명', '종목', 'stock_name', 'name'])
    code_col = find_col(df, ['종목코드', '코드', 'stock_code', 'ticker', 'code'])
    price_col = find_col(df, ['현재가', '종가', '가격', '현재가격', 'price', 'close'])
    score_col = find_col(df, ['점수', '종합점수', '실전점수', 'score'])
    sector_col = find_col(df, ['섹터', '분야', 'sector', 'theme'])
    action_col = find_col(df, ['진입', '판정', '신호', '매수', 'action', 'signal'])
    heat_col = find_col(df, ['과열', 'overheat'])
    source_col = find_col(df, ['후보출처', '출처', 'source'])
    memo_col = find_col(df, ['메모', '요약', '사유', 'guide', 'memo'])

    out = []
    for idx, (_, r) in enumerate(df.head(max_rows).iterrows(), 1):
        name = clean_text(r.get(name_col, '')) if name_col else ''
        code = normalize_code(r.get(code_col, '')) if code_col else ''
        if not name and not code:
            continue
        price = to_number(r.get(price_col, '')) if price_col else None
        score = to_number(r.get(score_col, '')) if score_col else None
        out.append({
            'rank': str(idx),
            'stock_name': name,
            'stock_code': code,
            'key': key_of(name, code),
            'current_price': fmt_int(price),
            'current_price_num': price,
            'score': fmt_int(score),
            'score_num': score,
            'sector': clean_text(r.get(sector_col, '')) if sector_col else '',
            'action': clean_text(r.get(action_col, '')) if action_col else '조건 확인 후 진입',
            'overheat': clean_text(r.get(heat_col, '')) if heat_col else '',
            'source': clean_text(r.get(source_col, '')) if source_col else '',
            'memo': clean_text(r.get(memo_col, '')) if memo_col else '',
        })
    return out


def build_entry_guides(cands: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    guides = []
    for c in cands[:15]:
        price = c.get('current_price_num')
        score = c.get('score_num')
        if price:
            entry_low = price * 0.98
            entry_high = price * 1.005
            target1 = price * 1.05
            target2 = price * 1.08
            stop = price * 0.96
            entry_guide = f'{fmt_int(entry_low)}~{fmt_int(entry_high)}원 구간에서 거래량/지지 확인 후 분할 진입'
            target_guide = f'1차 {fmt_int(target1)}원(+5%), 2차 {fmt_int(target2)}원(+8%) 부근 분할 익절 검토'
            stop_guide = f'{fmt_int(stop)}원(-4%) 이탈 또는 장대음봉 동반 시 손절/비중축소 검토'
        else:
            entry_guide = '현재가 미확인: 장중 현재가와 거래량 확인 후 판단'
            target_guide = '현재가 확인 후 +5%/+8% 기준으로 설정'
            stop_guide = '현재가 확인 후 -4% 또는 최근 지지선 이탈 기준으로 설정'
        if score is not None and score >= 90:
            priority = 'A: 우선 감시'
        elif score is not None and score >= 80:
            priority = 'B: 조건부 감시'
        else:
            priority = 'C: 관찰'
        guides.append({
            'rank': c.get('rank', ''),
            'stock_name': c.get('stock_name', ''),
            'stock_code': c.get('stock_code', ''),
            'sector': c.get('sector', ''),
            'score': c.get('score', ''),
            'current_price': c.get('current_price', ''),
            'priority': priority,
            'entry_action': c.get('action', '') or '조건 확인 후 진입',
            'entry_guide': entry_guide,
            'take_profit_guide': target_guide,
            'stop_loss_guide': stop_guide,
            'do_not_chase': '시초 급등/뉴스성 갭상승은 5~30분 눌림 확인 전 추격매수 금지',
            'check_points': '거래량 증가, 전일 고점 돌파/지지, 뉴스 지속성, 과열 여부 확인',
        })
    return guides


def candidate_price_map(cands: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    m = {}
    for c in cands:
        k = c.get('key') or key_of(c.get('stock_name'), c.get('stock_code'))
        if k:
            m[k] = c
    return m


def build_holding_judgments(cands: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    holdings = read_first_existing_csv(HOLDINGS_FILES)
    pmap = candidate_price_map(cands)
    if holdings.empty:
        return [{
            'stock_name': '', 'stock_code': '', 'quantity': '', 'avg_price': '', 'current_price': '',
            'return_pct': '', 'decision': 'INPUT_REQUIRED', 'guide': 'holdings_manual_input.csv를 입력하면 보유종목 판단이 표시됩니다.',
            'target_price': '', 'stop_loss': '', 'memo': ''
        }]
    rows = []
    for _, r in holdings.iterrows():
        rd = {clean_text(k): v for k, v in r.to_dict().items()}
        name = row_pick(rd, ['stock_name', '종목명', '종목'])
        code = normalize_code(row_pick(rd, ['stock_code', '종목코드', '코드']))
        if not name and not code:
            continue
        key = key_of(name, code)
        qty = to_number(row_pick(rd, ['quantity', '보유수량', '수량']))
        avg = to_number(row_pick(rd, ['avg_price', '평균단가', '평단', '매수단가']))
        target = to_number(row_pick(rd, ['target_price', '목표가', '익절가']))
        stop = to_number(row_pick(rd, ['stop_loss', '손절가', '손절']))
        matched = pmap.get(key, {})
        price = matched.get('current_price_num') or to_number(row_pick(rd, ['current_price', '현재가']))
        rate = ((price - avg) / avg * 100) if price and avg else None

        decision = 'NO_ACTION'
        guide = '현재가/평균단가 확인 후 관망'
        if price and stop and price <= stop:
            decision = 'STOP_WATCH'
            guide = '손절가 근접/이탈. 반등 실패 시 비중축소 또는 손절 검토'
        elif rate is not None and rate <= -5:
            decision = 'STOP_WATCH'
            guide = '손실률 -5% 부근. 손절 기준과 뉴스/거래량 재확인'
        elif price and target and price >= target:
            decision = 'TAKE_PROFIT_1'
            guide = '목표가 도달. 일부 익절 후 잔량 추세 추적 검토'
        elif rate is not None and rate >= 8:
            decision = 'TAKE_PROFIT_2'
            guide = '수익률 +8% 이상. 2차 익절/트레일링 스탑 검토'
        elif rate is not None and rate >= 4:
            decision = 'TAKE_PROFIT_1'
            guide = '수익권. 거래량 둔화 또는 저항 접근 시 일부 익절 검토'
        elif matched:
            decision = 'HOLD'
            guide = '오늘 후보에도 포함. 추세 유지 시 보유, 장중 거래량 확인'
        elif rate is not None and -3 < rate < 4:
            decision = 'HOLD'
            guide = '큰 변동 없음. 목표/손절 기준 유지'

        rows.append({
            'status': row_pick(rd, ['status', '상태']) or 'holding',
            'stock_name': name,
            'stock_code': code,
            'quantity': fmt_int(qty),
            'avg_price': fmt_int(avg),
            'current_price': fmt_int(price),
            'return_pct': fmt_pct(rate),
            'decision': decision,
            'guide': guide,
            'target_price': fmt_int(target),
            'stop_loss': fmt_int(stop),
            'in_today_candidates': 'Y' if matched else 'N',
            'candidate_rank': matched.get('rank', ''),
            'candidate_score': matched.get('score', ''),
            'memo': row_pick(rd, ['memo', '메모', 'note']),
        })
    return rows


def read_tracking() -> List[Dict[str, str]]:
    if not TRACKING_CSV.exists():
        return []
    with TRACKING_CSV.open('r', encoding='utf-8-sig', newline='') as f:
        return list(csv.DictReader(f))


def write_rows(path: Path, rows: List[Dict[str, Any]], headers: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
        w.writeheader()
        for r in rows:
            w.writerow({h: clean_text(r.get(h, '')) for h in headers})
    print(f'✅ CSV saved: {path} ({len(rows)} rows)')


def build_performance_tracking(cands: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    DATA.mkdir(parents=True, exist_ok=True)
    today = now_kst().strftime('%Y-%m-%d')
    ts = now_kst().strftime('%Y-%m-%d %H:%M:%S')
    sess = session_label()
    tracking = read_tracking()
    existing_ids = {r.get('recommendation_id', '') for r in tracking}

    for c in cands[:15]:
        rid = f"{today}_{sess}_{c.get('rank','')}_{c.get('stock_code') or c.get('stock_name')}"
        if rid in existing_ids:
            continue
        base_price = to_number(c.get('current_price'))
        tracking.append({
            'recommendation_id': rid,
            'recommend_date': today,
            'session': sess,
            'rank': c.get('rank', ''),
            'stock_name': c.get('stock_name', ''),
            'stock_code': c.get('stock_code', ''),
            'sector': c.get('sector', ''),
            'score': c.get('score', ''),
            'recommend_price': fmt_int(base_price),
            'latest_price': fmt_int(base_price),
            'latest_update_kst': ts,
            'age_days': '0',
            'latest_return_pct': '0.00%' if base_price else '',
            'd1_return_pct': '',
            'd3_return_pct': '',
            'd5_return_pct': '',
            'max_observed_return_pct': '0.00%' if base_price else '',
            'min_observed_return_pct': '0.00%' if base_price else '',
            'status': 'tracking',
        })

    pmap = candidate_price_map(cands)
    for r in tracking:
        k = key_of(r.get('stock_name', ''), r.get('stock_code', ''))
        match = pmap.get(k)
        rec_price = to_number(r.get('recommend_price'))
        if match and rec_price:
            latest_price = match.get('current_price_num')
            if latest_price:
                ret = (latest_price - rec_price) / rec_price * 100
                r['latest_price'] = fmt_int(latest_price)
                r['latest_update_kst'] = ts
                r['latest_return_pct'] = fmt_pct(ret)
                old_max = to_number(r.get('max_observed_return_pct'))
                old_min = to_number(r.get('min_observed_return_pct'))
                r['max_observed_return_pct'] = fmt_pct(max(ret, old_max if old_max is not None else ret))
                r['min_observed_return_pct'] = fmt_pct(min(ret, old_min if old_min is not None else ret))
                try:
                    age = (now_kst().date() - datetime.strptime(r.get('recommend_date',''), '%Y-%m-%d').date()).days
                except Exception:
                    age = 0
                r['age_days'] = str(age)
                if age >= 1 and not r.get('d1_return_pct'):
                    r['d1_return_pct'] = fmt_pct(ret)
                if age >= 3 and not r.get('d3_return_pct'):
                    r['d3_return_pct'] = fmt_pct(ret)
                if age >= 5 and not r.get('d5_return_pct'):
                    r['d5_return_pct'] = fmt_pct(ret)
                if age >= 5:
                    r['status'] = 'complete_5d'
                elif age >= 1:
                    r['status'] = 'tracking_updated'

    headers = [
        'recommendation_id', 'recommend_date', 'session', 'rank', 'stock_name', 'stock_code', 'sector', 'score',
        'recommend_price', 'latest_price', 'latest_update_kst', 'age_days', 'latest_return_pct',
        'd1_return_pct', 'd3_return_pct', 'd5_return_pct', 'max_observed_return_pct', 'min_observed_return_pct', 'status'
    ]
    write_rows(TRACKING_CSV, tracking, headers)
    latest = sorted(tracking, key=lambda r: (r.get('recommend_date',''), r.get('session',''), r.get('rank','')), reverse=True)[:80]
    write_rows(LATEST_PERF_CSV, latest, headers)
    return latest


def write_excel_sheet(wb, name: str, rows: List[Dict[str, Any]], headers: List[str]) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append([clean_text(r.get(h, '')) for h in headers])

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for idx, h in enumerate(headers, 1):
        width = min(max(len(h) + 4, 12), 38)
        if h in {'entry_guide', 'take_profit_guide', 'stop_loss_guide', 'guide', 'check_points', 'do_not_chase', 'memo'}:
            width = 42
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def update_excel(xlsx: Optional[Path], perf: List[Dict[str, str]], holdings: List[Dict[str, str]], guides: List[Dict[str, str]]) -> None:
    if xlsx is None or not xlsx.exists():
        print('⚠️ Excel file not found. Skip v11 Excel sheet update.')
        return
    wb = load_workbook(xlsx)
    perf_headers = ['recommend_date','session','rank','stock_name','stock_code','sector','score','recommend_price','latest_price','age_days','latest_return_pct','d1_return_pct','d3_return_pct','d5_return_pct','max_observed_return_pct','min_observed_return_pct','status']
    holding_headers = ['status','stock_name','stock_code','quantity','avg_price','current_price','return_pct','decision','guide','target_price','stop_loss','in_today_candidates','candidate_rank','candidate_score','memo']
    guide_headers = ['rank','stock_name','stock_code','sector','score','current_price','priority','entry_action','entry_guide','take_profit_guide','stop_loss_guide','do_not_chase','check_points']
    write_excel_sheet(wb, '추천성과_검증', perf, perf_headers)
    write_excel_sheet(wb, '보유종목_판단', holdings, holding_headers)
    write_excel_sheet(wb, '진입가이드_요약', guides, guide_headers)
    if '검증결과' in wb.sheetnames:
        ws = wb['검증결과']
    else:
        ws = wb.create_sheet('검증결과')
        ws.append(['검증항목','상태','메모'])
    ws.append(['v11.0 추천성과/보유판단 시트 생성', 'PASS', f'성과 {len(perf)}건 / 보유판단 {len(holdings)}건 / 진입가이드 {len(guides)}건'])
    wb.save(xlsx)
    print(f'✅ v11 Excel sheets updated: {xlsx}')


def badge_class(decision: str) -> str:
    d = (decision or '').upper()
    if 'STOP' in d or 'RISK' in d:
        return 'danger'
    if 'TAKE' in d:
        return 'profit'
    if 'ADD' in d:
        return 'watch'
    if 'HOLD' in d:
        return 'hold'
    return 'neutral'


def html_table(rows: List[Dict[str, Any]], headers: List[str], limit: int = 10) -> str:
    if not rows:
        return '<p class="v11-empty">표시할 데이터가 없습니다.</p>'
    th = ''.join(f'<th>{html.escape(h)}</th>' for h in headers)
    trs = []
    for r in rows[:limit]:
        tds = []
        for h in headers:
            val = clean_text(r.get(h, ''))
            if h in {'decision','priority','status'}:
                cls = badge_class(val)
                tds.append(f'<td><span class="v11-badge {cls}">{html.escape(val)}</span></td>')
            else:
                tds.append(f'<td>{html.escape(val)}</td>')
        trs.append('<tr>' + ''.join(tds) + '</tr>')
    return f'<div class="v11-table-wrap"><table class="v11-table"><thead><tr>{th}</tr></thead><tbody>{"".join(trs)}</tbody></table></div>'


def build_v11_html(perf: List[Dict[str, str]], holdings: List[Dict[str, str]], guides: List[Dict[str, str]]) -> str:
    updated = now_kst().strftime('%Y-%m-%d %H:%M:%S KST')
    css = """
<style>
.v11-section{margin:28px 0;padding:22px;border-radius:24px;background:#f8fafc;border:1px solid #e5e7eb;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;color:#111827}
.v11-section h2{margin:0 0 8px;font-size:24px}.v11-section p{line-height:1.55;color:#4b5563}.v11-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px;margin:16px 0}.v11-kpi{background:white;border-radius:18px;padding:16px;border:1px solid #e5e7eb}.v11-kpi b{display:block;font-size:20px;margin-top:6px}.v11-table-wrap{overflow-x:auto;background:white;border-radius:18px;border:1px solid #e5e7eb;margin:12px 0 22px}.v11-table{border-collapse:collapse;width:100%;font-size:14px}.v11-table th{background:#1f4e78;color:white;text-align:left;padding:10px;white-space:nowrap}.v11-table td{border-top:1px solid #e5e7eb;padding:10px;vertical-align:top}.v11-badge{display:inline-block;border-radius:999px;padding:5px 10px;font-weight:700;font-size:12px;white-space:nowrap}.v11-badge.hold{background:#dcfce7;color:#166534}.v11-badge.profit{background:#dbeafe;color:#1d4ed8}.v11-badge.watch{background:#fef9c3;color:#854d0e}.v11-badge.danger{background:#fee2e2;color:#b91c1c}.v11-badge.neutral{background:#e5e7eb;color:#374151}.v11-note{font-size:13px;color:#6b7280}.v11-empty{padding:12px;background:white;border-radius:12px}
</style>
"""
    kpi = f"""
<div class="v11-grid">
  <div class="v11-kpi"><span>보유종목 판단</span><b>{len(holdings)}건</b></div>
  <div class="v11-kpi"><span>진입/익절/손절 가이드</span><b>{len(guides)}건</b></div>
  <div class="v11-kpi"><span>성과 추적 누적</span><b>{len(perf)}건 표시</b></div>
</div>
"""
    holding_table = html_table(holdings[:8], ['stock_name','quantity','avg_price','current_price','return_pct','decision','guide'], 8)
    guide_table = html_table(guides[:8], ['rank','stock_name','score','current_price','priority','entry_guide','take_profit_guide','stop_loss_guide'], 8)
    perf_table = html_table(perf[:10], ['recommend_date','session','rank','stock_name','recommend_price','latest_price','latest_return_pct','d1_return_pct','d3_return_pct','d5_return_pct','status'], 10)
    return f"""<!-- V11_DASHBOARD_START -->
{css}
<section class="v11-section" id="v11-dashboard">
  <h2>v11.0 추천성과 검증 + 보유종목 판단</h2>
  <p>업데이트: {html.escape(updated)} / 이 섹션은 매수 확정 신호가 아니라, 후보 압축·보유 대응·위험관리 판단 보조용입니다.</p>
  {kpi}
  <h3>내 보유종목 판단</h3>
  {holding_table}
  <h3>진입·익절·손절 가이드</h3>
  {guide_table}
  <h3>추천성과 추적</h3>
  {perf_table}
  <p class="v11-note">성과 검증은 이후 리포트에서 같은 종목의 최신 가격이 확인될 때 갱신됩니다. 1일/3일/5일 수익률은 데이터가 누적될수록 채워집니다.</p>
</section>
<!-- V11_DASHBOARD_END -->"""


def inject_v11_into_html(path: Path, section: str) -> None:
    if not path.exists():
        return
    text = path.read_text(encoding='utf-8', errors='ignore')
    text = re.sub(r'<!-- V11_DASHBOARD_START -->.*?<!-- V11_DASHBOARD_END -->', '', text, flags=re.S)
    if '</body>' in text:
        text = text.replace('</body>', section + '\n</body>')
    else:
        text += '\n' + section
    path.write_text(text, encoding='utf-8')
    print(f'✅ v11 HTML injected: {path}')


def write_standalone_dashboard(perf: List[Dict[str,str]], holdings: List[Dict[str,str]], guides: List[Dict[str,str]]) -> None:
    V11_PAGE.mkdir(parents=True, exist_ok=True)
    section = build_v11_html(perf, holdings, guides)
    page = f"""<!doctype html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>v11 Dashboard</title></head>
<body style="margin:0;background:#eef2f7;padding:16px;">{section}</body></html>"""
    (V11_PAGE / 'index.html').write_text(page, encoding='utf-8')
    print('✅ standalone v11 dashboard created: docs/v11_dashboard/index.html')


def main() -> int:
    DATA.mkdir(parents=True, exist_ok=True)
    DOCS.mkdir(parents=True, exist_ok=True)
    xlsx = find_latest_xlsx()
    print(f'🚀 v11.0 dashboard start / session={session_label()} / xlsx={xlsx}')
    cands = candidate_rows(xlsx, max_rows=30)
    guides = build_entry_guides(cands)
    holdings = build_holding_judgments(cands)
    perf = build_performance_tracking(cands)

    guide_headers = ['rank','stock_name','stock_code','sector','score','current_price','priority','entry_action','entry_guide','take_profit_guide','stop_loss_guide','do_not_chase','check_points']
    holding_headers = ['status','stock_name','stock_code','quantity','avg_price','current_price','return_pct','decision','guide','target_price','stop_loss','in_today_candidates','candidate_rank','candidate_score','memo']
    write_rows(LATEST_ENTRY_GUIDE_CSV, guides, guide_headers)
    write_rows(LATEST_HOLDING_JUDGMENT_CSV, holdings, holding_headers)

    update_excel(xlsx, perf, holdings, guides)
    section = build_v11_html(perf, holdings, guides)
    inject_v11_into_html(DOCS / 'index.html', section)
    inject_v11_into_html(DOCS / 'latest' / 'index.html', section)
    write_standalone_dashboard(perf, holdings, guides)
    print('✅ v11.0 performance + holdings dashboard completed')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
