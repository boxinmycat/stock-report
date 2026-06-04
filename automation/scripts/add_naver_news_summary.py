#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Naver News + optional OpenAI interpretation v10.6
- Fetches Naver News API results.
- Writes 네이버뉴스_요약 and 네이버뉴스_상세 sheets into the latest workbook.
- If ENABLE_OPENAI_NEWS_ANALYSIS=true and OPENAI_API_KEY is present, adds OpenAI interpretation.
"""

from __future__ import annotations

import html
import json
import os
import re
import time
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd


def _find_latest_xlsx() -> Path:
    candidates = []
    for pattern in ["docs/reports/**/20*.xlsx", "20*.xlsx", "stock_report/**/*.xlsx"]:
        candidates.extend(Path(".").glob(pattern))
    candidates = [p for p in candidates if p.is_file() and not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("xlsx 리포트 파일을 찾지 못했습니다.")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _clean_html_text(s: str) -> str:
    s = html.unescape(str(s or ""))
    s = re.sub(r"<[^>]+>", "", s)
    s = s.replace("&quot;", '"')
    return re.sub(r"\s+", " ", s).strip()


def _read_top_context(xlsx_path: Path):
    top_names = []
    sectors = []
    try:
        df = pd.read_excel(xlsx_path, sheet_name="TOP후보_요약", engine="openpyxl")
        for col in df.columns:
            if "종목명" in str(col):
                top_names = [str(x).strip() for x in df[col].dropna().head(8).tolist()]
            if "섹터" in str(col) or "분야" in str(col):
                sectors = [str(x).strip() for x in df[col].dropna().head(8).tolist()]
    except Exception:
        pass
    sectors = [s for s, _ in Counter(sectors).most_common(5) if s and s.lower() != "nan"]
    return top_names, sectors


def _naver_search(query: str, display: int = 5):
    cid = os.getenv("NAVER_CLIENT_ID", "").strip()
    sec = os.getenv("NAVER_CLIENT_SECRET", "").strip()
    if not cid or not sec:
        return [{"query": query, "title": "네이버 뉴스 API 키 설정 필요", "description": "GitHub Secrets에 NAVER_CLIENT_ID / NAVER_CLIENT_SECRET을 설정하면 자동 수집됩니다.", "pubDate": "", "link": ""}]

    params = urllib.parse.urlencode({"query": query, "display": display, "sort": "date"})
    url = f"https://openapi.naver.com/v1/search/news.json?{params}"
    req = urllib.request.Request(url)
    req.add_header("X-Naver-Client-Id", cid)
    req.add_header("X-Naver-Client-Secret", sec)

    try:
        with urllib.request.urlopen(req, timeout=15) as res:
            data = json.loads(res.read().decode("utf-8"))
        rows = []
        for item in data.get("items", []):
            rows.append({
                "query": query,
                "title": _clean_html_text(item.get("title", "")),
                "description": _clean_html_text(item.get("description", "")),
                "pubDate": item.get("pubDate", ""),
                "link": item.get("originallink") or item.get("link", ""),
            })
        return rows
    except Exception as e:
        return [{"query": query, "title": "네이버 뉴스 수집 실패", "description": str(e), "pubDate": "", "link": ""}]


def _rule_sentiment(text: str) -> str:
    pos = ["상승", "강세", "호조", "수혜", "실적", "계약", "증가", "흑자", "개선", "확대", "신고가"]
    neg = ["하락", "약세", "우려", "손실", "적자", "감소", "리스크", "조사", "제재", "급락", "부진"]
    p = sum(w in text for w in pos)
    n = sum(w in text for w in neg)
    if p > n:
        return "긍정/호재성"
    if n > p:
        return "주의/리스크"
    return "중립/확인필요"


def _openai_interpret(context: str) -> str:
    if os.getenv("ENABLE_OPENAI_NEWS_ANALYSIS", "").lower() != "true":
        return ""
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        return ""
    model = os.getenv("OPENAI_NEWS_MODEL", "gpt-5.4-mini").strip() or "gpt-5.4-mini"
    prompt = (
        "다음은 한국 주식 자동 리포트용 네이버 뉴스 제목/요약이다. "
        "투자 조언을 단정하지 말고, 장전/장마감 브리핑 관점에서 핵심 이슈, 긍정 요인, 리스크, 확인할 조건을 한국어로 5줄 이내로 요약해줘.\n\n"
        + context[:12000]
    )
    payload = {
        "model": model,
        "input": prompt,
        "max_output_tokens": 700,
    }
    req = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as res:
            data = json.loads(res.read().decode("utf-8"))
        if data.get("output_text"):
            return data["output_text"].strip()
        parts = []
        for out in data.get("output", []):
            for content in out.get("content", []):
                if content.get("text"):
                    parts.append(content["text"])
        return "\n".join(parts).strip()
    except Exception as e:
        return f"OpenAI 해석 실패: {e}"


def _write_sheet(wb, name: str, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 22
    return ws


def add_news_summary():
    xlsx_path = _find_latest_xlsx()
    top_names, sectors = _read_top_context(xlsx_path)

    mode = os.getenv("REPORT_RUN_MODE", "")
    mode_query = "장전 증시 전망" if mode == "AM_PREMARKET" else "장마감 증시"
    queries = [
        f"{mode_query} 코스피 코스닥",
        "국내 증시 수급 외국인 기관",
    ]
    queries += [f"{s} 관련주 뉴스" for s in sectors[:3]]
    queries += [f"{n} 주가 뉴스" for n in top_names[:5]]

    detail_rows = []
    seen = set()
    for q in queries:
        for row in _naver_search(q, display=5):
            key = row.get("link") or (row.get("title"), row.get("description"))
            if key in seen:
                continue
            seen.add(key)
            text = f"{row.get('title','')} {row.get('description','')}"
            row["분위기"] = _rule_sentiment(text)
            detail_rows.append(row)
        time.sleep(0.1)

    by_group = {}
    for row in detail_rows:
        q = row.get("query", "")
        group = "시장뉴스"
        if "관련주" in q:
            group = "섹터뉴스"
        if "주가 뉴스" in q:
            group = "종목뉴스"
        by_group.setdefault(group, []).append(row)

    summary_rows = []
    for group, rows in by_group.items():
        titles = [r.get("title", "") for r in rows[:5] if r.get("title")]
        mood = Counter([r.get("분위기", "중립/확인필요") for r in rows]).most_common(1)[0][0] if rows else "중립/확인필요"
        summary_rows.append({
            "구분": group,
            "요약": " / ".join(titles[:3]) if titles else "수집된 뉴스가 없습니다.",
            "분위기": mood,
            "확인포인트": "뉴스 제목만으로 단정하지 말고 거래량·수급·공시·장중 가격 반응을 함께 확인",
        })

    context = "\n".join([f"- [{r.get('query')}] {r.get('title')} :: {r.get('description')}" for r in detail_rows[:30]])
    ai_text = _openai_interpret(context)
    if ai_text:
        summary_rows.insert(0, {
            "구분": "OpenAI 해석",
            "요약": ai_text,
            "분위기": "AI 보조해석",
            "확인포인트": "비용이 발생할 수 있는 OpenAI API 기반 요약. 투자 판단 확정이 아니라 보조 해석입니다.",
        })

    wb = openpyxl.load_workbook(xlsx_path)
    _write_sheet(wb, "네이버뉴스_요약", ["구분", "요약", "분위기", "확인포인트"], summary_rows)
    _write_sheet(wb, "네이버뉴스_상세", ["query", "title", "description", "pubDate", "link", "분위기"], detail_rows)
    wb.save(xlsx_path)
    print(f"✅ 네이버뉴스 요약 시트 생성 완료: {xlsx_path} / 상세 {len(detail_rows)}건 / OpenAI={'ON' if ai_text else 'OFF'}")


if __name__ == "__main__":
    add_news_summary()
