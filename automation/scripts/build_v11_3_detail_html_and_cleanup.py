#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
v11.3 Detail HTML + Workbook Slimming
- Creates mobile-friendly detail HTML pages for important data moved/hidden from Excel.
- Adds report basis info and TP/SL strategy guide sheets.
- Hides bulky diagnostic/detail sheets without deleting them.
- Exports detail CSV files under docs/data for Google Sheets.
"""
from __future__ import annotations

import html
import os
import re
import shutil
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KST = timezone(timedelta(hours=9))
ROOT = Path('.')
DOCS = ROOT / 'docs'
DETAILS = DOCS / 'details'
DATA = DOCS / 'data'
MOBILE = DOCS / 'mobile'

REPO_PAGES_BASE = os.getenv('PAGES_BASE_URL', 'https://boxinmycat.github.io/stock-report').rstrip('/')

CORE_VISIBLE_SHEETS = [
    '오늘_요약',
    '리포트_기준정보',
    '추천TOP_통합',
    '진입가이드_요약',
    '보유종목_판단',
    '보유종목_심화분석',
    '보유대응_가이드',
    '익절손절_전략기준',
    '익절손절_검증',
    '추천성과_검증',
    '네이버뉴스_요약',
    'HTML_상세보기_안내',
]

SHEET_CANDIDATES = {
    'news_summary': ['네이버뉴스_요약', 'Naver_News_Summary'],
    'news_detail': ['네이버뉴스_상세', 'Naver_News_Detail'],
    'candidates': ['추천TOP_통합', 'TOP후보_요약', '추천 리스트', '추천리스트'],
    'continuous': ['연속추천_관찰'],
    'performance': ['추천성과_검증', 'recommendation_tracking'],
    'holding_judgment': ['보유종목_판단', '보유종목_심화분석'],
    'entry_exit': ['진입가이드_요약', '보유대응_가이드'],
    'validation': ['검증결과', 'Validation'],
}

TPSL_STRATEGIES = [
    {
        '전략': '안정형',
        '1차익절': '+6%',
        '2차익절': '+10%',
        '기본손절': '-5%',
        '최대보유': '5거래일',
        '적용대상': '단기 후보 / 변동성 큰 종목',
        '운영메모': '빠른 회전용. 수익을 짧게 확인하되 과도한 잦은 매매는 주의',
    },
    {
        '전략': '습관형',
        '1차익절': '+8%',
        '2차익절': '+15%',
        '기본손절': '-7%',
        '최대보유': '10거래일',
        '적용대상': '기본값 / 소액 분할투자',
        '운영메모': '소액 투자 습관형 기본값. 손절폭은 넓히되 종목당 비중은 작게 유지',
    },
    {
        '전략': '추세형',
        '1차익절': '+10%',
        '2차익절': '+20%',
        '기본손절': '-10%',
        '최대보유': '15~20거래일',
        '적용대상': '강한 추세 / 연속추천 후보',
        '운영메모': '강한 종목을 길게 보는 기준. 급등 추격이 아니라 눌림 확인 후 적용',
    },
]

CSS = """
:root{--bg:#f6f7fb;--card:#fff;--ink:#182230;--muted:#667085;--line:#e5e7eb;--blue:#1f4e78;--green:#157347;--red:#b42318;--yellow:#b54708;}
*{box-sizing:border-box} body{margin:0;background:var(--bg);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,'Noto Sans KR',Arial,sans-serif;color:var(--ink);line-height:1.45}
a{color:#0b5cad;text-decoration:none} a:hover{text-decoration:underline}
.wrap{max-width:1080px;margin:0 auto;padding:18px 14px 42px}.hero{background:linear-gradient(135deg,#16233f,#1f4e78);color:#fff;border-radius:18px;padding:22px;margin-bottom:14px;box-shadow:0 10px 28px rgba(16,24,40,.15)}
.hero h1{margin:0 0 8px;font-size:25px}.hero p{margin:4px 0;color:#e5eefc}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px}.card{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:15px;box-shadow:0 4px 14px rgba(16,24,40,.06);margin-bottom:12px}.card h2{font-size:19px;margin:0 0 10px}.card h3{font-size:16px;margin:4px 0 8px}.muted{color:var(--muted);font-size:13px}.badge{display:inline-flex;align-items:center;padding:4px 9px;border-radius:999px;background:#eef4ff;color:#194185;font-size:12px;font-weight:700;margin:2px}.badge.green{background:#ecfdf3;color:#027a48}.badge.red{background:#fef3f2;color:#b42318}.badge.yellow{background:#fffaeb;color:#b54708}.badge.gray{background:#f2f4f7;color:#475467}
.table-wrap{width:100%;overflow:auto;border:1px solid var(--line);border-radius:12px;background:#fff}table{border-collapse:collapse;width:100%;min-width:760px}th,td{border-bottom:1px solid var(--line);padding:9px 10px;text-align:left;vertical-align:top;font-size:13px}th{background:#f2f4f7;font-weight:800;white-space:nowrap}tr:hover td{background:#fafafa}.small-table table{min-width:420px}
.details-list details{background:#fff;border:1px solid var(--line);border-radius:14px;margin:10px 0;overflow:hidden}summary{cursor:pointer;padding:13px 14px;font-weight:800;background:#fbfcff}.details-body{padding:13px 14px;border-top:1px solid var(--line)}
.news-card{border:1px solid var(--line);border-radius:14px;padding:13px;margin:10px 0;background:#fff}.news-title{font-weight:800;font-size:15px}.news-desc{color:#344054;margin:7px 0}.pill-row{margin:7px 0}.footer{color:var(--muted);font-size:12px;text-align:center;margin-top:30px}.nav{display:flex;flex-wrap:wrap;gap:8px;margin-top:14px}.nav a{background:#fff;color:#1f4e78;border:1px solid rgba(255,255,255,.7);border-radius:999px;padding:8px 11px;font-weight:700}.warning{background:#fffaeb;border:1px solid #fedf89;color:#7a2e0e;border-radius:14px;padding:12px;margin:12px 0}.ok{background:#ecfdf3;border:1px solid #abefc6;color:#054f31;border-radius:14px;padding:12px;margin:12px 0}
@media(max-width:640px){.wrap{padding:12px 10px 28px}.hero{border-radius:15px;padding:17px}.hero h1{font-size:21px}.card{border-radius:14px;padding:13px}th,td{font-size:12px;padding:8px}table{min-width:680px}.grid{grid-template-columns:1fr}}
"""


def now_kst() -> datetime:
    return datetime.now(KST)


def session_name(dt: Optional[datetime] = None) -> str:
    dt = dt or now_kst()
    if dt.hour < 12:
        return 'AM 장전'
    return 'PM 장마감'


def session_data_basis(sess: str) -> str:
    if sess.startswith('AM'):
        return '장전 리포트: 가격/거래량은 주로 전일 종가·전일 거래량 기준, 뉴스는 생성시각까지 검색 기준'
    return '장마감 리포트: 데이터 소스 반영 상태에 따라 당일 종가·거래량 중심, 뉴스는 생성시각까지 검색 기준'


def clean_cell(value) -> str:
    if pd.isna(value):
        return ''
    text = str(value)
    text = text.replace('\r', ' ').replace('\n', ' ').strip()
    text = re.sub(r'\s+', ' ', text)
    # Remove common pandas leaked artifacts
    if text.lower() in {'nan', 'none', 'nat'}:
        return ''
    if 'dtype: object' in text and 'Name:' in text:
        text = re.sub(r'Name:\s*\d+\s*,\s*dtype:\s*object', '', text).strip()
    text = text.replace('dtype: object', '').strip()
    return text


def clean_df(df: pd.DataFrame, max_rows: Optional[int] = None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out.columns = [clean_cell(c) or f'col_{i+1}' for i, c in enumerate(out.columns)]
    for c in out.columns:
        out[c] = out[c].map(clean_cell)
    out = out.dropna(how='all')
    out = out.loc[:, [c for c in out.columns if c and not str(c).startswith('Unnamed')]]
    if max_rows is not None:
        out = out.head(max_rows)
    return out


def find_latest_xlsx() -> Optional[Path]:
    candidates: List[Path] = []
    patterns = ['20*.xlsx', '*report*.xlsx', 'stock_report/**/*.xlsx']
    for pattern in patterns:
        candidates.extend([p for p in ROOT.glob(pattern) if p.is_file() and not p.name.startswith('~$')])
    if not candidates:
        return None
    return max(set(candidates), key=lambda p: p.stat().st_mtime)


def read_sheet(xlsx: Optional[Path], sheet_names: Iterable[str], max_rows: Optional[int] = None) -> pd.DataFrame:
    if xlsx is None or not xlsx.exists():
        return pd.DataFrame()
    try:
        xl = pd.ExcelFile(xlsx)
        available = xl.sheet_names
        for name in sheet_names:
            if name in available:
                return clean_df(pd.read_excel(xlsx, sheet_name=name), max_rows=max_rows)
        # fuzzy contains
        for name in sheet_names:
            for av in available:
                if str(name).replace(' ', '') in str(av).replace(' ', ''):
                    return clean_df(pd.read_excel(xlsx, sheet_name=av), max_rows=max_rows)
    except Exception as e:
        print(f'⚠️ 시트 읽기 실패: {sheet_names} / {e}')
    return pd.DataFrame()


def read_csv_candidates(paths: Iterable[Path], max_rows: Optional[int] = None) -> pd.DataFrame:
    for p in paths:
        if p.exists():
            for enc in ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr']:
                try:
                    return clean_df(pd.read_csv(p, encoding=enc), max_rows=max_rows)
                except Exception:
                    pass
    return pd.DataFrame()


def df_to_table(df: pd.DataFrame, empty_msg: str = '표시할 데이터가 없습니다.', max_rows: Optional[int] = None) -> str:
    df = clean_df(df, max_rows=max_rows)
    if df.empty:
        return f'<div class="warning">{html.escape(empty_msg)}</div>'
    headers = ''.join(f'<th>{html.escape(str(c))}</th>' for c in df.columns)
    rows = []
    for _, row in df.iterrows():
        tds = ''.join(f'<td>{html.escape(clean_cell(row.get(c, "")))}</td>' for c in df.columns)
        rows.append(f'<tr>{tds}</tr>')
    return '<div class="table-wrap"><table><thead><tr>'+headers+'</tr></thead><tbody>'+''.join(rows)+'</tbody></table></div>'


def write_html(path: Path, title: str, body: str, subtitle: str = '') -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    nav = f"""
    <div class="nav">
      <a href="{REPO_PAGES_BASE}/mobile/">모바일 홈</a>
      <a href="{REPO_PAGES_BASE}/latest/">최신 리포트</a>
      <a href="{REPO_PAGES_BASE}/v11_dashboard/">v11 대시보드</a>
      <a href="{REPO_PAGES_BASE}/v11_holdings/">보유종목 심화</a>
      <a href="{REPO_PAGES_BASE}/details/">상세 데이터</a>
    </div>
    """
    doc = f"""<!doctype html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(title)}</title><style>{CSS}</style></head><body><div class="wrap">
<section class="hero"><h1>{html.escape(title)}</h1><p>{html.escape(subtitle)}</p>{nav}</section>
{body}
<div class="footer">Generated by stock-report v11.3 detail module · {now_kst().strftime('%Y-%m-%d %H:%M:%S KST')}</div>
</div></body></html>"""
    path.write_text(doc, encoding='utf-8')


def make_news_detail_page(news_detail: pd.DataFrame, news_summary: pd.DataFrame) -> str:
    parts = []
    parts.append('<div class="card"><h2>네이버뉴스 상세</h2><p class="muted">엑셀에서는 숨기거나 뒤로 밀 수 있는 기사 상세 목록을 모바일에서 보기 좋게 카드형으로 정리한 페이지입니다.</p></div>')
    if not news_summary.empty:
        parts.append('<div class="card"><h2>뉴스 요약</h2>' + df_to_table(news_summary, max_rows=30) + '</div>')
    if news_detail.empty:
        parts.append('<div class="warning">네이버뉴스 상세 데이터가 아직 없습니다. NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 또는 뉴스 생성 단계를 확인하세요.</div>')
        return ''.join(parts)
    cols = list(news_detail.columns)
    title_col = next((c for c in cols if any(k in str(c).lower() for k in ['title','제목','뉴스제목'])), cols[0])
    desc_col = next((c for c in cols if any(k in str(c).lower() for k in ['description','desc','요약','내용'])), None)
    link_col = next((c for c in cols if any(k in str(c).lower() for k in ['link','url','링크'])), None)
    group_col = next((c for c in cols if any(k in str(c).lower() for k in ['category','구분','type','query','검색어','종목명'])), None)
    detail = clean_df(news_detail, max_rows=120)
    if group_col:
        grouped = detail.groupby(group_col, dropna=False)
        for g, gdf in grouped:
            gname = clean_cell(g) or '기타 뉴스'
            parts.append(f'<div class="card"><h2>{html.escape(gname)}</h2>')
            for _, row in gdf.head(20).iterrows():
                title = clean_cell(row.get(title_col, '제목 없음')) or '제목 없음'
                desc = clean_cell(row.get(desc_col, '')) if desc_col else ''
                link = clean_cell(row.get(link_col, '')) if link_col else ''
                badges = ''.join(f'<span class="badge gray">{html.escape(str(c))}: {html.escape(clean_cell(row.get(c, ""))[:40])}</span>' for c in cols[:4] if c not in [title_col, desc_col, link_col])
                title_html = f'<a href="{html.escape(link)}" target="_blank" rel="noopener">{html.escape(title)}</a>' if link else html.escape(title)
                parts.append(f'<div class="news-card"><div class="news-title">{title_html}</div><div class="pill-row">{badges}</div><div class="news-desc">{html.escape(desc)}</div></div>')
            parts.append('</div>')
    else:
        parts.append('<div class="card"><h2>전체 뉴스</h2>')
        for _, row in detail.iterrows():
            title = clean_cell(row.get(title_col, '제목 없음')) or '제목 없음'
            desc = clean_cell(row.get(desc_col, '')) if desc_col else ''
            link = clean_cell(row.get(link_col, '')) if link_col else ''
            title_html = f'<a href="{html.escape(link)}" target="_blank" rel="noopener">{html.escape(title)}</a>' if link else html.escape(title)
            parts.append(f'<div class="news-card"><div class="news-title">{title_html}</div><div class="news-desc">{html.escape(desc)}</div></div>')
        parts.append('</div>')
    return ''.join(parts)


def export_csv(df: pd.DataFrame, filename: str) -> None:
    DATA.mkdir(parents=True, exist_ok=True)
    clean_df(df).to_csv(DATA / filename, index=False, encoding='utf-8-sig')


def style_ws(ws) -> None:
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='DDDDDD')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for col_idx in range(1, min(ws.max_column, 12) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


def set_sheet_from_df(wb, sheet_name: str, df: pd.DataFrame, index: int = 1) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, index=min(index, len(wb.sheetnames)))
    df = clean_df(df)
    if df.empty:
        df = pd.DataFrame([{'메모': '표시할 데이터가 없습니다.'}])
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([clean_cell(row.get(c, '')) for c in df.columns])
    style_ws(ws)


def update_workbook(xlsx: Optional[Path], report_basis_df: pd.DataFrame, tpsl_df: pd.DataFrame) -> None:
    if xlsx is None or not xlsx.exists():
        print('⚠️ 엑셀 파일이 없어 엑셀 경량화/시트 추가는 건너뜁니다.')
        return
    wb = load_workbook(xlsx)
    set_sheet_from_df(wb, '리포트_기준정보', report_basis_df, index=1)
    set_sheet_from_df(wb, '익절손절_전략기준', tpsl_df, index=7)
    guide_df = pd.DataFrame([
        {'구분': '모바일 상세 데이터', '주소': f'{REPO_PAGES_BASE}/details/', '설명': '엑셀에서 숨긴 상세 뉴스/원본/검증 로그를 모바일에서 확인'},
        {'구분': '네이버뉴스 상세', '주소': f'{REPO_PAGES_BASE}/details/naver_news.html', '설명': '기사 상세 목록과 링크'},
        {'구분': '후보 상세', '주소': f'{REPO_PAGES_BASE}/details/candidate_detail.html', '설명': '추천 후보 상세 원본'},
        {'구분': '연속추천 상세', '주소': f'{REPO_PAGES_BASE}/details/continuous.html', '설명': '연속추천 관찰 원본'},
        {'구분': '백테스트/검증 로그', '주소': f'{REPO_PAGES_BASE}/details/diagnostics.html', '설명': '백테스트 요약 및 실행 검증 로그'},
    ])
    set_sheet_from_df(wb, 'HTML_상세보기_안내', guide_df, index=2)

    # Hide bulky sheets but never delete them
    for ws in wb.worksheets:
        name = ws.title
        should_hide = False
        if name in ['네이버뉴스_상세', '검증결과']:
            should_hide = True
        if any(key in name.upper() for key in ['RAW', 'DETAIL']) and name not in CORE_VISIBLE_SHEETS:
            should_hide = True
        if '상세' in name and name not in ['보유종목_심화분석', 'HTML_상세보기_안내']:
            should_hide = True
        if should_hide and len([s for s in wb.worksheets if s.sheet_state == 'visible']) > 1:
            ws.sheet_state = 'hidden'

    # Move visible core sheets to front when they exist
    for target_idx, sheet_name in enumerate(reversed(CORE_VISIBLE_SHEETS)):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb._sheets.remove(ws)
            wb._sheets.insert(0, ws)
    wb.save(xlsx)
    print(f'✅ 엑셀 경량화/기준정보 추가 완료: {xlsx}')


def main() -> int:
    DETAILS.mkdir(parents=True, exist_ok=True)
    DATA.mkdir(parents=True, exist_ok=True)
    MOBILE.mkdir(parents=True, exist_ok=True)

    dt = now_kst()
    sess = session_name(dt)
    xlsx = find_latest_xlsx()
    print(f'🔎 v11.3 대상 엑셀: {xlsx if xlsx else "없음"}')

    report_basis_df = pd.DataFrame([
        {'항목': '리포트 생성시각', '내용': dt.strftime('%Y-%m-%d %H:%M:%S KST')},
        {'항목': '리포트 구분', '내용': sess},
        {'항목': '가격/거래량 기준', '내용': session_data_basis(sess)},
        {'항목': '뉴스 수집 기준', '내용': '생성시각까지 네이버뉴스 API 검색 결과 기준'},
        {'항목': '보유종목 기준', '내용': '저장소 루트 holdings_manual_input.csv 기준'},
        {'항목': '매매기록 기준', '내용': '저장소 루트 trade_log_manual_input.csv 기준'},
        {'항목': '주의', '내용': '리포트는 판단 보조용이며 실제 매수/매도 전 현재가·거래량·호가 확인 필요'},
    ])
    tpsl_df = pd.DataFrame(TPSL_STRATEGIES)

    data: Dict[str, pd.DataFrame] = {}
    for key, sheets in SHEET_CANDIDATES.items():
        max_rows = 150 if key == 'news_detail' else 80
        data[key] = read_sheet(xlsx, sheets, max_rows=max_rows)

    # Fallbacks from generated csv files
    if data['news_detail'].empty:
        data['news_detail'] = read_csv_candidates([DATA / 'latest_news_detail.csv', DATA / 'latest_news_summary.csv'], max_rows=150)
    if data['candidates'].empty:
        data['candidates'] = read_csv_candidates([DATA / 'latest_candidates.csv', DATA / 'latest_candidate_detail.csv'], max_rows=100)
    if data['holding_judgment'].empty:
        data['holding_judgment'] = read_csv_candidates([DATA / 'latest_holding_deep_analysis.csv', DATA / 'latest_holding_judgment.csv', DATA / 'latest_holdings.csv'], max_rows=100)
    if data['entry_exit'].empty:
        data['entry_exit'] = read_csv_candidates([DATA / 'latest_holding_action_guide.csv', DATA / 'latest_entry_exit_guide.csv'], max_rows=100)

    export_csv(report_basis_df, 'latest_report_basis.csv')
    export_csv(tpsl_df, 'latest_tpsl_strategy_guide.csv')
    export_csv(data['news_detail'], 'latest_news_detail.csv')
    export_csv(data['candidates'], 'latest_candidate_detail.csv')
    export_csv(data['continuous'], 'latest_continuous_detail.csv')
    export_csv(data['validation'], 'latest_run_log.csv')

    update_workbook(xlsx, report_basis_df, tpsl_df)

    # HTML pages
    index_body = f"""
    <div class="grid">
      <div class="card"><h2>네이버뉴스 상세</h2><p>숨김 처리된 기사 상세와 링크를 모바일 카드로 확인합니다.</p><a class="badge green" href="{REPO_PAGES_BASE}/details/naver_news.html">열기</a></div>
      <div class="card"><h2>후보 상세 데이터</h2><p>추천 후보 원본/상세표를 확인합니다.</p><a class="badge" href="{REPO_PAGES_BASE}/details/candidate_detail.html">열기</a></div>
      <div class="card"><h2>연속추천 상세</h2><p>연속추천 관찰 원본을 확인합니다.</p><a class="badge" href="{REPO_PAGES_BASE}/details/continuous.html">열기</a></div>
      <div class="card"><h2>보유/매도 가이드</h2><p>보유종목 대응, 진입·익절·손절 기준을 확인합니다.</p><a class="badge yellow" href="{REPO_PAGES_BASE}/details/holding_action.html">열기</a></div>
      <div class="card"><h2>기준정보·익절손절</h2><p>데이터 기준시각과 전략별 익절/손절 기본값입니다.</p><a class="badge" href="{REPO_PAGES_BASE}/details/risk_rules.html">열기</a></div>
      <div class="card"><h2>진단/검증 로그</h2><p>실행 검증 로그와 백테스트 요약을 확인합니다.</p><a class="badge gray" href="{REPO_PAGES_BASE}/details/diagnostics.html">열기</a></div>
    </div>
    """
    write_html(DETAILS / 'index.html', '상세 데이터 센터', index_body, '엑셀에서는 가볍게 숨기고, 모바일에서는 자세히 보는 영역')

    write_html(DETAILS / 'naver_news.html', '네이버뉴스 상세', make_news_detail_page(data['news_detail'], data['news_summary']), '시장·섹터·종목별 기사 상세와 링크')

    candidate_body = '<div class="card"><h2>추천 후보 상세</h2><p class="muted">엑셀에서 통합/축약되는 후보 상세 데이터를 모바일에서 확인합니다.</p></div>' + df_to_table(data['candidates'], '추천 후보 상세 데이터가 없습니다.', max_rows=120)
    write_html(DETAILS / 'candidate_detail.html', '추천 후보 상세', candidate_body, '추천TOP/후보 원본 상세')

    continuous_body = '<div class="card"><h2>연속추천 상세</h2><p class="muted">반복 등장 후보는 중요한 신호이므로, 엑셀에서는 요약하고 상세는 모바일에서 확인합니다.</p></div>' + df_to_table(data['continuous'], '연속추천 데이터가 없습니다.', max_rows=100)
    write_html(DETAILS / 'continuous.html', '연속추천 상세', continuous_body, '연속추천 관찰 데이터')

    holding_body = '<div class="card"><h2>보유종목 대응</h2><p class="muted">보유종목 판단, 진입/추가매수/익절/손절 가이드를 함께 확인합니다.</p></div>'
    holding_body += '<div class="card"><h2>보유종목 판단</h2>' + df_to_table(data['holding_judgment'], '보유종목 판단 데이터가 없습니다.', max_rows=80) + '</div>'
    holding_body += '<div class="card"><h2>보유대응/진입가이드</h2>' + df_to_table(data['entry_exit'], '진입/익절/손절 가이드 데이터가 없습니다.', max_rows=80) + '</div>'
    write_html(DETAILS / 'holding_action.html', '보유종목 대응·매도 가이드', holding_body, 'HOLD / 익절 / 손절 / 추가매수 관찰')

    risk_body = '<div class="card"><h2>리포트 기준정보</h2>' + df_to_table(report_basis_df, max_rows=20) + '</div>'
    risk_body += '<div class="card"><h2>익절/손절 전략 기본값</h2>' + df_to_table(tpsl_df, max_rows=20) + '</div>'
    risk_body += '<div class="warning"><b>주의:</b> 아직 검증중인 기본값입니다. 추천성과가 쌓이면 AM/PM, 신규/보유, 과열/정상 기준으로 실제 유효성을 비교해야 합니다.</div>'
    write_html(DETAILS / 'risk_rules.html', '기준정보·익절손절 전략', risk_body, '데이터 기준시각과 TP/SL 전략 기본값')

    diag_body = '<div class="card"><h2>추천성과/검증 로그</h2>' + df_to_table(data['performance'], '추천성과 데이터가 아직 없습니다.', max_rows=80) + '</div>'
    diag_body += '<div class="card"><h2>실행 검증 로그</h2>' + df_to_table(data['validation'], '검증 로그 데이터가 없습니다.', max_rows=80) + '</div>'
    write_html(DETAILS / 'diagnostics.html', '진단·검증 로그', diag_body, '숨김 처리되는 실행 로그와 성과 추적 데이터')

    mobile_body = f"""
    <div class="grid">
      <div class="card"><h2>최신 리포트</h2><p>오늘 최신 HTML 리포트</p><a class="badge green" href="{REPO_PAGES_BASE}/latest/">열기</a></div>
      <div class="card"><h2>v11 대시보드</h2><p>추천성과·보유판단</p><a class="badge" href="{REPO_PAGES_BASE}/v11_dashboard/">열기</a></div>
      <div class="card"><h2>보유종목 심화</h2><p>보유종목 대응/매도 가이드</p><a class="badge yellow" href="{REPO_PAGES_BASE}/v11_holdings/">열기</a></div>
      <div class="card"><h2>상세 데이터 센터</h2><p>뉴스 상세, 원본표, 검증 로그</p><a class="badge" href="{REPO_PAGES_BASE}/details/">열기</a></div>
      <div class="card"><h2>구글시트용 데이터</h2><p>latest_*.csv 자동 연동용</p><a class="badge gray" href="{REPO_PAGES_BASE}/data/latest_report_basis.csv">기준정보 CSV</a></div>
    </div>
    """
    write_html(MOBILE / 'index.html', '주식 리포트 모바일 홈', mobile_body, '최신 리포트와 상세 데이터를 핸드폰에서 빠르게 확인')

    print('✅ v11.3 상세 HTML/엑셀 경량화 완료')
    print(f'   - {REPO_PAGES_BASE}/details/')
    print(f'   - {REPO_PAGES_BASE}/details/naver_news.html')
    print(f'   - {REPO_PAGES_BASE}/mobile/')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
