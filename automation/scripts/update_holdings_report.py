#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Holdings / trade-log post processor v10.7
- Reads 보유종목_수동입력.csv and 매매기록_수동입력.csv from repository root.
- Adds 보유종목_관리 and 매매기록_관리 sheets into the latest xlsx report.
- Matches holdings against TOP후보_요약 / 연속추천_관찰 to show whether current holdings are also current candidates.

This script never places orders. It only adds management information to the report.
"""

from __future__ import annotations

import math
import os
import re
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HOLDINGS_FILE = os.getenv("HOLDINGS_CSV_FILE", "보유종목_수동입력.csv")
TRADE_LOG_FILE = os.getenv("TRADE_LOG_CSV_FILE", "매매기록_수동입력.csv")


def _find_latest_xlsx() -> Path:
    candidates = []
    for pattern in ["docs/reports/**/20*.xlsx", "20*.xlsx", "stock_report/**/*.xlsx"]:
        candidates.extend(Path(".").glob(pattern))
    candidates = [p for p in candidates if p.is_file() and not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("xlsx 리포트 파일을 찾지 못했습니다.")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _clean_text(v) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return str(v).strip()


def _norm_code(v) -> str:
    s = _clean_text(v)
    s = re.sub(r"[^0-9]", "", s)
    return s.zfill(6) if s else ""


def _to_number(v) -> Optional[float]:
    s = _clean_text(v)
    if not s:
        return None
    s = s.replace(",", "").replace("원", "").replace("%", "")
    try:
        x = float(s)
        if math.isnan(x):
            return None
        return x
    except Exception:
        return None


def _fmt_num(v) -> str:
    x = _to_number(v)
    if x is None:
        return ""
    if abs(x - round(x)) < 1e-9:
        return f"{int(round(x)):,}"
    return f"{x:,.2f}".rstrip("0").rstrip(".")


def _read_csv(path: str | Path) -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        return pd.DataFrame()
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        try:
            df = pd.read_csv(p, encoding=enc)
            df.columns = [_clean_text(c) for c in df.columns]
            return df.dropna(how="all")
        except Exception:
            continue
    print(f"⚠️ CSV 읽기 실패: {p}")
    return pd.DataFrame()


def _find_col(df: pd.DataFrame, candidates) -> Optional[str]:
    if df is None or df.empty:
        return None
    for c in df.columns:
        cs = str(c).replace(" ", "")
        for cand in candidates:
            if cand.replace(" ", "") in cs:
                return c
    return None


def _read_sheet_safe(xlsx: Path, sheet_name: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(xlsx, sheet_name=sheet_name, engine="openpyxl")
        df.columns = [_clean_text(c) for c in df.columns]
        return df.dropna(how="all")
    except Exception:
        return pd.DataFrame()


def _make_key(name, code) -> str:
    code = _norm_code(code)
    if code:
        return code
    return _clean_text(name).replace(" ", "")


def _build_candidate_map(xlsx: Path):
    candidate_map = {}
    for sheet, label in [("TOP후보_요약", "TOP후보"), ("연속추천_관찰", "연속추천")]:
        df = _read_sheet_safe(xlsx, sheet)
        if df.empty:
            continue
        name_col = _find_col(df, ["종목명", "종목"])
        code_col = _find_col(df, ["종목코드", "코드"])
        price_col = _find_col(df, ["현재가", "종가", "가격"])
        score_col = _find_col(df, ["실전점수", "점수"])
        entry_col = _find_col(df, ["진입판정", "판정"])
        sector_col = _find_col(df, ["섹터/분야", "섹터", "분야"])
        for _, row in df.iterrows():
            key = _make_key(row.get(name_col, "") if name_col else "", row.get(code_col, "") if code_col else "")
            if not key:
                continue
            old = candidate_map.get(key, {})
            status = old.get("리포트상태", "")
            if label not in status:
                status = (status + "+" + label).strip("+") if status else label
            candidate_map[key] = {
                **old,
                "리포트상태": status,
                "리포트현재가": row.get(price_col, "") if price_col else old.get("리포트현재가", ""),
                "리포트점수": row.get(score_col, "") if score_col else old.get("리포트점수", ""),
                "리포트진입판정": row.get(entry_col, "") if entry_col else old.get("리포트진입판정", ""),
                "리포트섹터": row.get(sector_col, "") if sector_col else old.get("리포트섹터", ""),
            }
    return candidate_map


def _normalize_holdings(holdings: pd.DataFrame, candidate_map: dict) -> pd.DataFrame:
    if holdings.empty:
        return pd.DataFrame([{
            "상태": "입력필요",
            "종목명": "",
            "종목코드": "",
            "보유수량": "",
            "평균단가": "",
            "현재가": "",
            "매입금액": "",
            "평가금액": "",
            "평가손익": "",
            "수익률(%)": "",
            "목표가": "",
            "손절가": "",
            "리포트상태": "보유종목_수동입력.csv를 입력하면 자동 표시됩니다.",
            "관리메모": "GitHub 저장소가 공개라면 수량/평균단가는 노출될 수 있으니 주의하세요.",
        }])

    name_col = _find_col(holdings, ["종목명", "종목"])
    code_col = _find_col(holdings, ["종목코드", "코드"])
    qty_col = _find_col(holdings, ["보유수량", "수량"])
    avg_col = _find_col(holdings, ["평균단가", "매수단가", "평단"])
    target_col = _find_col(holdings, ["목표가", "익절가"])
    stop_col = _find_col(holdings, ["손절가", "손절"])
    status_col = _find_col(holdings, ["상태"])
    strategy_col = _find_col(holdings, ["전략구분", "전략"])
    memo_col = _find_col(holdings, ["메모", "매수이유"])
    buy_date_col = _find_col(holdings, ["매수일", "진입일"])

    rows = []
    for _, row in holdings.iterrows():
        name = _clean_text(row.get(name_col, "")) if name_col else ""
        code = _norm_code(row.get(code_col, "")) if code_col else ""
        if not name and not code:
            continue
        key = _make_key(name, code)
        match = candidate_map.get(key, {})
        qty = _to_number(row.get(qty_col, "")) if qty_col else None
        avg = _to_number(row.get(avg_col, "")) if avg_col else None
        price = _to_number(match.get("리포트현재가", ""))
        target = _to_number(row.get(target_col, "")) if target_col else None
        stop = _to_number(row.get(stop_col, "")) if stop_col else None
        buy_amt = qty * avg if qty is not None and avg is not None else None
        eval_amt = qty * price if qty is not None and price is not None else None
        pnl = eval_amt - buy_amt if eval_amt is not None and buy_amt is not None else None
        pnl_rate = (pnl / buy_amt * 100) if pnl is not None and buy_amt else None

        notes = []
        if match.get("리포트상태"):
            notes.append("오늘 리포트 후보/관찰군에 포함")
        else:
            notes.append("오늘 후보에는 미포함")
        if price is not None and target is not None and price >= target:
            notes.append("목표가 도달/분할익절 검토")
        if price is not None and stop is not None and price <= stop:
            notes.append("손절가 접근/리스크 확인")
        if price is None:
            notes.append("현재가 매칭 안 됨: 종목코드 확인")

        rows.append({
            "상태": _clean_text(row.get(status_col, "보유중")) if status_col else "보유중",
            "종목명": name,
            "종목코드": code,
            "보유수량": _fmt_num(qty),
            "평균단가": _fmt_num(avg),
            "현재가": _fmt_num(price),
            "매입금액": _fmt_num(buy_amt),
            "평가금액": _fmt_num(eval_amt),
            "평가손익": _fmt_num(pnl),
            "수익률(%)": _fmt_num(pnl_rate),
            "목표가": _fmt_num(target),
            "손절가": _fmt_num(stop),
            "매수일": _clean_text(row.get(buy_date_col, "")) if buy_date_col else "",
            "전략구분": _clean_text(row.get(strategy_col, "")) if strategy_col else "",
            "리포트상태": match.get("리포트상태", "보유관리"),
            "리포트점수": _fmt_num(match.get("리포트점수", "")),
            "리포트진입판정": _clean_text(match.get("리포트진입판정", "")),
            "섹터/분야": _clean_text(match.get("리포트섹터", "")),
            "관리메모": " / ".join(notes),
            "사용자메모": _clean_text(row.get(memo_col, "")) if memo_col else "",
        })

    return pd.DataFrame(rows)


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 30) + 1)]
        width = min(max(max((len(v) for v in values), default=8) + 2, 10), 42)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def _write_df_sheet(wb, name: str, df: pd.DataFrame):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 2)
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([row.get(c, "") for c in df.columns])
    _style_sheet(ws)


def update_holdings_report():
    xlsx = _find_latest_xlsx()
    candidate_map = _build_candidate_map(xlsx)
    holdings = _read_csv(HOLDINGS_FILE)
    trades = _read_csv(TRADE_LOG_FILE)
    holdings_out = _normalize_holdings(holdings, candidate_map)

    wb = openpyxl.load_workbook(xlsx)
    _write_df_sheet(wb, "보유종목_관리", holdings_out)
    if not trades.empty:
        _write_df_sheet(wb, "매매기록_관리", trades.fillna(""))
    else:
        _write_df_sheet(wb, "매매기록_관리", pd.DataFrame([{
            "메모": "매매기록_수동입력.csv를 입력하면 매수/매도 기록이 이 시트에 표시됩니다."
        }]))
    wb.save(xlsx)
    print(f"✅ 보유종목 관리 시트 생성 완료: {xlsx} / 보유 {len(holdings_out)}건 / 거래기록 {len(trades)}건")


if __name__ == "__main__":
    update_holdings_report()
