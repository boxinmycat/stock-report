#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import html
import math
import os
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
except Exception:
    load_workbook = None

KST = timezone(timedelta(hours=9))
ROOT = Path('.')
DOCS = ROOT / 'docs'
DATA_DIR = DOCS / 'data'
STRATEGY_DIR = DOCS / 'strategy'
TRACKING_PATH = DATA_DIR / 'recommendation_tracking.csv'

STRATEGIES = [
    {'strategy_code': 'SAFE', 'strategy_name': '안정형', 'tp1_pct': 0.06, 'tp2_pct': 0.10, 'sl_pct': -0.05, 'max_hold_days': 5},
    {'strategy_code': 'HABIT', 'strategy_name': '습관형', 'tp1_pct': 0.08, 'tp2_pct': 0.15, 'sl_pct': -0.07, 'max_hold_days': 10},
    {'strategy_code': 'TREND', 'strategy_name': '추세형', 'tp1_pct': 0.10, 'tp2_pct': 0.20, 'sl_pct': -0.10, 'max_hold_days': 20},
]

CANDIDATE_SHEETS = ['추천TOP_통합', 'TOP후보_요약', '추천 리스트', '추천리스트', '진입가이드_요약']
PRICE_COL_CANDIDATES = ['추천가', '기준가', '현재가', '현재가격', '종가', 'close', 'price', '현재 주가', '현재주가', 'last_price', 'entry_price']
NAME_COL_CANDIDATES = ['종목명', '종목', 'stock_name', 'name', 'Name']
CODE_COL_CANDIDATES = ['종목코드', '코드', 'stock_code', 'code', 'Code']
SCORE_COL_CANDIDATES = ['추천점수', '점수', 'score', 'Score', '종합점수']
SOURCE_COL_CANDIDATES = ['후보출처', '출처', 'source', 'Source', '스크리너명']
SECTOR_COL_CANDIDATES = ['분야', '섹터', '테마', 'sector', 'theme', 'industry']
RANK_COL_CANDIDATES = ['순위', 'rank', 'Rank']


def now_kst() -> datetime:
    return datetime.now(KST)


def today_key(dt: Optional[datetime] = None) -> str:
    return (dt or now_kst()).strftime('%Y%m%d')


def detect_session(dt: Optional[datetime] = None) -> str:
    env_session = os.environ.get('REPORT_SESSION') or os.environ.get('SESSION_NAME')
    if env_session:
        s = str(env_session).upper()
        if 'PM' in s:
            return 'PM'
        if 'AM' in s:
            return 'AM'
    dt = dt or now_kst()
    return 'AM' if dt.hour < 12 else 'PM'


def norm_text(x) -> str:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ''
    s = str(x).strip()
    if s.lower() in {'nan', 'none', 'null'}:
        return ''
    return s


def norm_code(x) -> str:
    s = norm_text(x)
    if not s:
        return ''
    s = re.sub(r'\.0$', '', s)
    if s.isdigit() and len(s) < 6:
        s = s.zfill(6)
    return s


def to_number(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        if pd.isna(x):
            return None
        return float(x)
    s = str(x).strip()
    if not s or s.lower() in {'nan', 'none', 'null', '-'}:
        return None
    s = s.replace(',', '').replace('원', '').replace('%', '').replace('+', '')
    s = re.sub(r'[^\d\.\-]', '', s)
    if not s or s in {'-', '.', '-.'}:
        return None
    try:
        return float(s)
    except Exception:
        return None


def find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = list(df.columns)
    compact = {str(c).replace(' ', '').lower(): c for c in cols}
    for cand in candidates:
        key = cand.replace(' ', '').lower()
        if key in compact:
            return compact[key]
    for cand in candidates:
        key = cand.replace(' ', '').lower()
        for c in cols:
            if key and key in str(c).replace(' ', '').lower():
                return c
    return None


def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy().dropna(how='all')
    df = df.loc[:, [c for c in df.columns if not str(c).startswith('Unnamed')]]
    return df


def find_latest_xlsx() -> Optional[Path]:
    candidates = []
    for pattern in ['20*.xlsx', '*.xlsx']:
        candidates.extend([p for p in ROOT.glob(pattern) if not p.name.startswith('~$')])
    if not candidates:
        candidates = [p for p in ROOT.rglob('*.xlsx') if not p.name.startswith('~$')]
    return max(candidates, key=lambda p: p.stat().st_mtime) if candidates else None


def make_key(name, code) -> str:
    code = norm_code(code)
    if code:
        return f'C:{code}'
    return f'N:{norm_text(name)}'


def normalize_candidate_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    name_col = find_col(df, NAME_COL_CANDIDATES)
    code_col = find_col(df, CODE_COL_CANDIDATES)
    price_col = find_col(df, PRICE_COL_CANDIDATES)
    score_col = find_col(df, SCORE_COL_CANDIDATES)
    source_col = find_col(df, SOURCE_COL_CANDIDATES)
    sector_col = find_col(df, SECTOR_COL_CANDIDATES)
    rank_col = find_col(df, RANK_COL_CANDIDATES)
    if not name_col and not code_col:
        return pd.DataFrame()
    rows = []
    for _, r in df.iterrows():
        name = norm_text(r.get(name_col, '')) if name_col else ''
        code = norm_code(r.get(code_col, '')) if code_col else ''
        if not name and not code:
            continue
        price = to_number(r.get(price_col, None)) if price_col else None
        score = to_number(r.get(score_col, None)) if score_col else None
        rank = to_number(r.get(rank_col, None)) if rank_col else None
        rows.append({
            'rank': int(rank) if rank is not None and rank >= 1 else len(rows) + 1,
            'stock_name': name,
            'stock_code': code,
            'snapshot_price': price,
            'score': score,
            'source': norm_text(r.get(source_col, '')) if source_col else '',
            'sector': norm_text(r.get(sector_col, '')) if sector_col else '',
            'source_sheet': norm_text(r.get('_source_sheet', '')),
            'recommendation_model': norm_text(r.get('recommendation_model', '')) or 'A_BALANCED',
            'recommendation_model_name': norm_text(r.get('recommendation_model_name', '')) or 'A_균형형',
        })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out['_key'] = out.apply(lambda x: make_key(x.get('stock_name'), x.get('stock_code')), axis=1)
    out = out.drop_duplicates('_key', keep='first').drop(columns=['_key'])
    out = out.sort_values('rank').head(int(os.environ.get('V11_TRACK_TOP_N', '20')))
    return out.reset_index(drop=True)


def read_candidates_from_xlsx(xlsx_path: Path) -> pd.DataFrame:
    frames = []
    try:
        xls = pd.ExcelFile(xlsx_path)
    except Exception:
        return pd.DataFrame()
    for sheet in CANDIDATE_SHEETS:
        if sheet not in xls.sheet_names:
            continue
        try:
            df = clean_df(pd.read_excel(xlsx_path, sheet_name=sheet))
            if not df.empty:
                df['_source_sheet'] = sheet
                frames.append(df)
        except Exception:
            pass
    if not frames:
        return pd.DataFrame()
    for pref in ['추천TOP_통합', 'TOP후보_요약', '추천 리스트']:
        for df in frames:
            try:
                if str(df['_source_sheet'].iloc[0]) == pref:
                    return normalize_candidate_df(df)
            except Exception:
                pass
    return normalize_candidate_df(frames[0])


def read_candidates_from_csv() -> pd.DataFrame:
    paths = [DATA_DIR / 'latest_candidates.csv', DATA_DIR / 'latest_entry_exit_guide.csv', DATA_DIR / 'latest_holding_judgment.csv']
    for path in paths:
        if not path.exists():
            continue
        for enc in ['utf-8-sig', 'utf-8', 'cp949']:
            try:
                df = clean_df(pd.read_csv(path, encoding=enc))
                if not df.empty:
                    return normalize_candidate_df(df)
            except Exception:
                pass
    return pd.DataFrame()


def load_tracking() -> pd.DataFrame:
    if TRACKING_PATH.exists():
        for enc in ['utf-8-sig', 'utf-8', 'cp949']:
            try:
                return pd.read_csv(TRACKING_PATH, encoding=enc)
            except Exception:
                pass
    return pd.DataFrame()


def save_csv(df: pd.DataFrame, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding='utf-8-sig')


def make_signal_id(date_key: str, session: str, rank: int, name: str, code: str, model: str) -> str:
    key = norm_code(code) or re.sub(r'\s+', '_', norm_text(name))
    return f'{date_key}_{session}_{model}_{rank:02d}_{key}'


def append_current_snapshot(tracking: pd.DataFrame, candidates: pd.DataFrame, dt: datetime) -> pd.DataFrame:
    if candidates is None or candidates.empty:
        return tracking
    date_key = today_key(dt)
    session = detect_session(dt)
    generated_at = dt.strftime('%Y-%m-%d %H:%M:%S KST')
    existing_ids = set(tracking['signal_id'].astype(str)) if not tracking.empty and 'signal_id' in tracking.columns else set()
    rows = []
    for _, r in candidates.iterrows():
        rank = int(r.get('rank') or len(rows) + 1)
        model = norm_text(r.get('recommendation_model')) or 'A_BALANCED'
        sid = make_signal_id(date_key, session, rank, r.get('stock_name', ''), r.get('stock_code', ''), model)
        if sid in existing_ids:
            continue
        entry_price = r.get('snapshot_price')
        rows.append({
            'signal_id': sid,
            'snapshot_date': date_key,
            'snapshot_time': generated_at,
            'session': session,
            'rank': rank,
            'recommendation_model': model,
            'recommendation_model_name': norm_text(r.get('recommendation_model_name')) or 'A_균형형',
            'stock_name': norm_text(r.get('stock_name')),
            'stock_code': norm_code(r.get('stock_code')),
            'entry_model': 'M0_REPORT_PRICE',
            'entry_price': entry_price,
            'score': r.get('score'),
            'source': norm_text(r.get('source')),
            'sector': norm_text(r.get('sector')),
            'observed_count': 0,
            'latest_observed_price': entry_price,
            'latest_observed_at': generated_at if entry_price else '',
            'max_observed_price': entry_price,
            'min_observed_price': entry_price,
            'max_return_pct': 0.0 if entry_price else None,
            'min_return_pct': 0.0 if entry_price else None,
            'latest_return_pct': 0.0 if entry_price else None,
            'days_elapsed': 0,
            'status': 'TRACKING',
            'note': '추천 스냅샷 자동 저장',
        })
    if rows:
        tracking = pd.concat([tracking, pd.DataFrame(rows)], ignore_index=True) if not tracking.empty else pd.DataFrame(rows)
    return tracking


def update_tracking_observations(tracking: pd.DataFrame, candidates: pd.DataFrame, dt: datetime) -> pd.DataFrame:
    if tracking is None or tracking.empty:
        return tracking
    if candidates is None:
        candidates = pd.DataFrame()
    price_map = {}
    for _, r in candidates.iterrows():
        price = r.get('snapshot_price')
        if price is None or pd.isna(price):
            continue
        price_map[make_key(r.get('stock_name'), r.get('stock_code'))] = float(price)
        if norm_text(r.get('stock_name')):
            price_map[f'N:{norm_text(r.get("stock_name"))}'] = float(price)
    generated_at = dt.strftime('%Y-%m-%d %H:%M:%S KST')
    out = tracking.copy()
    for idx, r in out.iterrows():
        entry = to_number(r.get('entry_price'))
        if entry is None or entry <= 0:
            continue
        key = make_key(r.get('stock_name'), r.get('stock_code'))
        obs = price_map.get(key) or price_map.get(f'N:{norm_text(r.get("stock_name"))}')
        if obs is None:
            continue
        prev_max = to_number(r.get('max_observed_price')) or entry
        prev_min = to_number(r.get('min_observed_price')) or entry
        max_p = max(prev_max, obs)
        min_p = min(prev_min, obs)
        out.at[idx, 'latest_observed_price'] = obs
        out.at[idx, 'latest_observed_at'] = generated_at
        out.at[idx, 'max_observed_price'] = max_p
        out.at[idx, 'min_observed_price'] = min_p
        out.at[idx, 'max_return_pct'] = round((max_p / entry - 1) * 100, 2)
        out.at[idx, 'min_return_pct'] = round((min_p / entry - 1) * 100, 2)
        out.at[idx, 'latest_return_pct'] = round((obs / entry - 1) * 100, 2)
        out.at[idx, 'observed_count'] = int(to_number(r.get('observed_count')) or 0) + 1
    for idx, r in out.iterrows():
        sd = norm_text(r.get('snapshot_date'))
        try:
            snap_dt = datetime.strptime(sd, '%Y%m%d').replace(tzinfo=KST)
            out.at[idx, 'days_elapsed'] = max(0, (dt.date() - snap_dt.date()).days)
        except Exception:
            pass
    return out


def evaluate_strategy_for_signal(row: pd.Series, strategy: Dict) -> Dict:
    entry = to_number(row.get('entry_price'))
    if entry is None or entry <= 0:
        return {'result': 'NO_PRICE', 'sim_return_pct': None, 'hit_tp1': False, 'hit_tp2': False, 'hit_sl': False, 'memo': '추천가/기준가 없음'}
    max_ret = to_number(row.get('max_return_pct'))
    min_ret = to_number(row.get('min_return_pct'))
    latest_ret = to_number(row.get('latest_return_pct'))
    days = int(to_number(row.get('days_elapsed')) or 0)
    max_ret = max_ret if max_ret is not None else 0.0
    min_ret = min_ret if min_ret is not None else 0.0
    latest_ret = latest_ret if latest_ret is not None else 0.0
    tp1 = strategy['tp1_pct'] * 100
    tp2 = strategy['tp2_pct'] * 100
    sl = strategy['sl_pct'] * 100
    max_days = int(strategy['max_hold_days'])
    hit_tp2 = max_ret >= tp2
    hit_tp1 = max_ret >= tp1
    hit_sl = min_ret <= sl
    if hit_sl and hit_tp1:
        result = 'BOTH_HIT_ORDER_UNKNOWN'; sim_return = sl; memo = '익절/손절 모두 관측. 선후 불명이라 보수적으로 손절 가정'
    elif hit_tp2:
        result = 'TP2_HIT'; sim_return = tp2; memo = '2차 익절 기준 도달'
    elif hit_tp1:
        result = 'TP1_HIT'; sim_return = tp1; memo = '1차 익절 기준 도달'
    elif hit_sl:
        result = 'SL_HIT'; sim_return = sl; memo = '손절 기준 도달'
    elif days >= max_days:
        result = 'TIME_EXIT'; sim_return = latest_ret; memo = f'최대보유 {max_days}일 경과. 최근 관측 수익률 기준'
    else:
        result = 'TRACKING'; sim_return = latest_ret; memo = '검증 진행 중'
    return {'result': result, 'sim_return_pct': round(sim_return, 2), 'hit_tp1': bool(hit_tp1), 'hit_tp2': bool(hit_tp2), 'hit_sl': bool(hit_sl), 'memo': memo}


def build_strategy_detail(tracking: pd.DataFrame) -> pd.DataFrame:
    if tracking is None or tracking.empty:
        return pd.DataFrame()
    rows = []
    for _, r in tracking.iterrows():
        for st in STRATEGIES:
            ev = evaluate_strategy_for_signal(r, st)
            rows.append({
                'signal_id': r.get('signal_id'), 'snapshot_date': r.get('snapshot_date'), 'session': r.get('session'),
                'recommendation_model': r.get('recommendation_model', 'A_BALANCED'), 'recommendation_model_name': r.get('recommendation_model_name', 'A_균형형'),
                'stock_name': r.get('stock_name'), 'stock_code': r.get('stock_code'), 'rank': r.get('rank'), 'entry_price': r.get('entry_price'),
                'score': r.get('score'), 'source': r.get('source'), 'sector': r.get('sector'),
                'strategy_code': st['strategy_code'], 'strategy_name': st['strategy_name'], 'tp1_pct': int(st['tp1_pct'] * 100), 'tp2_pct': int(st['tp2_pct'] * 100), 'sl_pct': int(st['sl_pct'] * 100), 'max_hold_days': st['max_hold_days'],
                'result': ev['result'], 'sim_return_pct': ev['sim_return_pct'], 'max_return_pct': r.get('max_return_pct'), 'min_return_pct': r.get('min_return_pct'), 'latest_return_pct': r.get('latest_return_pct'), 'days_elapsed': r.get('days_elapsed'),
                'hit_tp1': ev['hit_tp1'], 'hit_tp2': ev['hit_tp2'], 'hit_sl': ev['hit_sl'], 'memo': ev['memo'],
            })
    return pd.DataFrame(rows)


def build_strategy_summary(detail: pd.DataFrame) -> pd.DataFrame:
    if detail is None or detail.empty:
        return pd.DataFrame()
    rows = []
    for keys, g in detail.groupby(['session', 'recommendation_model', 'recommendation_model_name', 'strategy_code', 'strategy_name'], dropna=False):
        session, model, model_name, strategy_code, strategy_name = keys
        signals = len(g)
        valid = g[g['result'] != 'NO_PRICE'].copy()
        closed = valid[valid['result'].isin(['TP1_HIT', 'TP2_HIT', 'SL_HIT', 'TIME_EXIT', 'BOTH_HIT_ORDER_UNKNOWN'])]
        returns = pd.to_numeric(valid['sim_return_pct'], errors='coerce').dropna()
        wins = returns[returns > 0]
        tp1_rate = valid['hit_tp1'].astype(bool).mean() * 100 if len(valid) else None
        sl_rate = valid['hit_sl'].astype(bool).mean() * 100 if len(valid) else None
        avg_ret = returns.mean() if len(returns) else None
        med_ret = returns.median() if len(returns) else None
        avg_max = pd.to_numeric(valid['max_return_pct'], errors='coerce').mean() if len(valid) else None
        avg_min = pd.to_numeric(valid['min_return_pct'], errors='coerce').mean() if len(valid) else None
        win_rate = len(wins) / len(returns) * 100 if len(returns) else None
        if signals < 50:
            status = '샘플부족'; memo = '최소 50개 이상부터 참고'
        elif avg_ret is not None and avg_ret > 1.0 and (sl_rate or 0) < 45:
            status = '우수후보'; memo = '평균수익률과 손절도달률이 비교적 양호'
        elif avg_ret is not None and avg_ret > 0:
            status = '검증중'; memo = '플러스 기대값이지만 추가 관찰 필요'
        else:
            status = '주의'; memo = '현재 관측 기준 기대값 낮음'
        rows.append({'session': session, 'recommendation_model': model, 'recommendation_model_name': model_name, 'strategy_code': strategy_code, 'strategy_name': strategy_name, 'signals': signals, 'valid_price_signals': len(valid), 'closed_or_hit': len(closed), 'win_rate_pct': round(win_rate, 2) if win_rate is not None else '', 'avg_return_pct': round(avg_ret, 2) if avg_ret is not None else '', 'median_return_pct': round(med_ret, 2) if med_ret is not None else '', 'tp1_hit_rate_pct': round(tp1_rate, 2) if tp1_rate is not None else '', 'sl_hit_rate_pct': round(sl_rate, 2) if sl_rate is not None else '', 'avg_max_return_pct': round(avg_max, 2) if avg_max is not None else '', 'avg_min_return_pct': round(avg_min, 2) if avg_min is not None else '', 'status': status, 'memo': memo})
    out = pd.DataFrame(rows)
    if not out.empty and 'avg_return_pct' in out.columns:
        out['_sort'] = pd.to_numeric(out['avg_return_pct'], errors='coerce')
        out = out.sort_values(['session', '_sort'], ascending=[True, False], na_position='last').drop(columns=['_sort'])
    return out


def build_model_summary(detail: pd.DataFrame) -> pd.DataFrame:
    if detail is None or detail.empty:
        return pd.DataFrame()
    habit = detail[detail['strategy_code'] == 'HABIT'].copy()
    if habit.empty:
        habit = detail.copy()
    rows = []
    for keys, g in habit.groupby(['recommendation_model', 'recommendation_model_name', 'session'], dropna=False):
        model, model_name, session = keys
        returns = pd.to_numeric(g['sim_return_pct'], errors='coerce').dropna()
        rows.append({'recommendation_model': model, 'recommendation_model_name': model_name, 'session': session, 'signals': len(g), 'avg_return_pct_habit': round(returns.mean(), 2) if len(returns) else '', 'win_rate_pct_habit': round(returns.gt(0).mean() * 100, 2) if len(returns) else '', 'tp1_hit_rate_pct_habit': round(g['hit_tp1'].astype(bool).mean() * 100, 2) if len(g) else '', 'sl_hit_rate_pct_habit': round(g['hit_sl'].astype(bool).mean() * 100, 2) if len(g) else '', 'status': '샘플부족' if len(g) < 50 else '검증가능', 'memo': '현재 추천조건은 A_균형형 기준모델. 이후 보수형/모멘텀형과 비교 예정'})
    return pd.DataFrame(rows)


def write_sheet(wb, sheet_name: str, df: pd.DataFrame, index: int = 0, hide: bool = False):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, index=min(index, len(wb.sheetnames)))
    if df is None or df.empty:
        df = pd.DataFrame([{'메모': '데이터가 아직 없습니다. 리포트가 누적되면 자동으로 채워집니다.'}])
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter][:60]:
            max_len = max(max_len, min(40, len(norm_text(cell.value)) + 2))
        ws.column_dimensions[letter].width = max_len
    ws.freeze_panes = 'A2'
    if hide:
        ws.sheet_state = 'hidden'


def update_workbook(xlsx_path: Optional[Path], tracking: pd.DataFrame, detail: pd.DataFrame, summary: pd.DataFrame, model_summary: pd.DataFrame):
    if load_workbook is None or xlsx_path is None or not xlsx_path.exists():
        return
    try:
        wb = load_workbook(xlsx_path)
        write_sheet(wb, '추천스냅샷_추적', tracking.tail(300), index=2)
        write_sheet(wb, '익절손절_검증', summary, index=3)
        write_sheet(wb, '추천조건_검증', model_summary, index=4)
        write_sheet(wb, '전략검증_상세', detail.tail(500), index=5, hide=True)
        wb.save(xlsx_path)
        print(f'✅ v11.4 엑셀 시트 반영 완료: {xlsx_path}')
    except Exception as e:
        print(f'⚠️ v11.4 엑셀 반영 실패: {e}')


def html_table(df: pd.DataFrame, max_rows: int = 30) -> str:
    if df is None or df.empty:
        return '<div class="empty">데이터가 아직 없습니다.</div>'
    show = df.head(max_rows).copy().fillna('')
    cols = list(show.columns)
    thead = ''.join(f'<th>{html.escape(str(c))}</th>' for c in cols)
    rows = []
    for _, r in show.iterrows():
        rows.append('<tr>' + ''.join(f'<td>{html.escape(norm_text(r.get(c, "")))}</td>' for c in cols) + '</tr>')
    return f'<div class="table-wrap"><table><thead><tr>{thead}</tr></thead><tbody>{"".join(rows)}</tbody></table></div>'


def render_strategy_html(tracking: pd.DataFrame, detail: pd.DataFrame, summary: pd.DataFrame, model_summary: pd.DataFrame):
    STRATEGY_DIR.mkdir(parents=True, exist_ok=True)
    now = now_kst().strftime('%Y-%m-%d %H:%M:%S KST')
    signal_count = len(tracking) if tracking is not None else 0
    valid_count = len(detail[detail['result'] != 'NO_PRICE']) if detail is not None and not detail.empty else 0
    best_line = '샘플이 누적되면 자동 표시됩니다.'
    if summary is not None and not summary.empty:
        s2 = summary.copy()
        s2['avg_sort'] = pd.to_numeric(s2['avg_return_pct'], errors='coerce')
        s2 = s2.sort_values('avg_sort', ascending=False, na_position='last')
        top = s2.iloc[0]
        best_line = f"{top.get('session','')} / {top.get('strategy_name','')} / 평균 {top.get('avg_return_pct','')}% / 상태 {top.get('status','')}"
    css = """
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f7fb;margin:0;color:#172033}.wrap{max-width:1180px;margin:0 auto;padding:22px}.hero{background:linear-gradient(135deg,#1f4e78,#182d4a);color:white;border-radius:22px;padding:24px;margin-bottom:18px;box-shadow:0 14px 35px rgba(20,40,80,.18)}.hero h1{margin:0 0 10px;font-size:26px}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px;margin:16px 0}.card{background:white;border-radius:18px;padding:18px;box-shadow:0 8px 24px rgba(20,40,80,.08);border:1px solid #e8edf5}.metric{font-size:28px;font-weight:800;color:#1f4e78}.label{font-size:13px;color:#667085;margin-top:4px}h2{font-size:20px;margin:26px 0 10px;color:#172033}.note{background:#fff7e6;border:1px solid #ffd591;border-radius:14px;padding:14px;margin:14px 0;color:#594214}.table-wrap{overflow:auto;background:white;border-radius:16px;border:1px solid #e8edf5;box-shadow:0 8px 24px rgba(20,40,80,.06)}table{border-collapse:collapse;width:100%;font-size:13px;min-width:860px}th{background:#1f4e78;color:white;text-align:left;padding:10px;position:sticky;top:0}td{border-bottom:1px solid #edf2f7;padding:9px;vertical-align:top}tr:nth-child(even) td{background:#fafcff}.pill{display:inline-block;background:#e8f2ff;color:#1f4e78;padding:4px 9px;border-radius:999px;font-size:12px;font-weight:700}.links a{display:inline-block;margin:5px 8px 5px 0;color:#1f4e78;text-decoration:none;font-weight:700}.empty{padding:16px;color:#667085}@media(max-width:680px){.wrap{padding:12px}.hero{padding:18px;border-radius:18px}table{font-size:12px}.metric{font-size:22px}}
"""
    recent_tracking = tracking
    if tracking is not None and not tracking.empty and 'snapshot_time' in tracking.columns:
        recent_tracking = tracking.sort_values('snapshot_time', ascending=False).head(50)
    recent_detail = detail
    if detail is not None and not detail.empty and 'snapshot_date' in detail.columns:
        recent_detail = detail.sort_values('snapshot_date', ascending=False).head(100)
    doc = f"""<!doctype html><html lang='ko'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>v11.4 추천전략 검증</title><style>{css}</style></head><body><div class='wrap'><div class='hero'><h1>v11.4 추천목록 자동검증 + 익절/손절 전략</h1><div>생성시각: {html.escape(now)}</div><div style='margin-top:8px'><span class='pill'>판단 보조용</span> <span class='pill'>A_균형형 기준모델</span> <span class='pill'>+8/+15/-7 습관형 검증</span></div></div><div class='grid'><div class='card'><div class='metric'>{signal_count}</div><div class='label'>누적 추천 신호</div></div><div class='card'><div class='metric'>{valid_count}</div><div class='label'>가격 검증 가능 신호</div></div><div class='card'><div class='metric'>3</div><div class='label'>동시 검증 전략 수</div></div><div class='card'><div class='metric'>검증중</div><div class='label'>{html.escape(best_line)}</div></div></div><div class='note'>이 화면은 실제 매매기록이 아니라 <b>리포트 추천목록 전체</b>를 기준으로 전략을 검증합니다. 실제 매매기록 검증은 이후 v11.5에서 별도로 붙이는 것이 좋습니다. 가격 데이터는 리포트 실행 시점에 관측된 가격을 누적해 계산하므로, 초반에는 샘플 부족으로 표시될 수 있습니다.</div><div class='links'><a href='../mobile/'>모바일 홈</a><a href='../latest/'>최신 리포트</a><a href='../v11_dashboard/'>v11 대시보드</a><a href='../details/'>상세 데이터 센터</a><a href='../data/latest_strategy_validation_summary.csv'>CSV 요약</a></div><h2>익절/손절 전략별 요약</h2>{html_table(summary, 50)}<h2>추천조건/모델 성과 요약</h2>{html_table(model_summary, 50)}<h2>최근 추천 스냅샷 추적</h2>{html_table(recent_tracking, 50)}<h2>전략 검증 상세 최근 100개</h2>{html_table(recent_detail, 100)}</div></body></html>"""
    (STRATEGY_DIR / 'index.html').write_text(doc, encoding='utf-8')
    print(f'✅ v11.4 HTML 생성 완료: {STRATEGY_DIR / "index.html"}')


def update_mobile_links():
    mobile = DOCS / 'mobile' / 'index.html'
    if not mobile.exists():
        return
    try:
        txt = mobile.read_text(encoding='utf-8')
        if 'strategy/' in txt or '추천전략 검증' in txt:
            return
        insert = "<a class='card' href='../strategy/'><h2>v11.4 추천전략 검증</h2><p>추천목록 전체를 기준으로 안정형/습관형/추세형 익절·손절 전략을 비교합니다.</p></a>"
        if '</main>' in txt:
            txt = txt.replace('</main>', insert + '\n</main>')
        elif '</body>' in txt:
            txt = txt.replace('</body>', insert + '\n</body>')
        else:
            txt += insert
        mobile.write_text(txt, encoding='utf-8')
        print('✅ 모바일 홈에 v11.4 전략 검증 링크 추가')
    except Exception as e:
        print(f'⚠️ 모바일 링크 추가 실패: {e}')


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    STRATEGY_DIR.mkdir(parents=True, exist_ok=True)
    dt = now_kst()
    xlsx = find_latest_xlsx()
    candidates = read_candidates_from_csv()
    if candidates.empty and xlsx:
        candidates = read_candidates_from_xlsx(xlsx)
    if candidates.empty:
        print('⚠️ v11.4: 추천 후보 데이터를 찾지 못했습니다. tracking만 유지합니다.')
    else:
        print(f'✅ v11.4: 추천 후보 {len(candidates)}개 인식')
    tracking = load_tracking()
    tracking = append_current_snapshot(tracking, candidates, dt)
    tracking = update_tracking_observations(tracking, candidates, dt)
    if not tracking.empty:
        if 'snapshot_time' in tracking.columns:
            tracking = tracking.sort_values('snapshot_time', ascending=False).head(1000).sort_values('snapshot_time')
        save_csv(tracking, TRACKING_PATH)
        save_csv(tracking.tail(300), DATA_DIR / 'latest_recommendation_tracking.csv')
    else:
        save_csv(pd.DataFrame(), TRACKING_PATH)
        save_csv(pd.DataFrame(), DATA_DIR / 'latest_recommendation_tracking.csv')
    detail = build_strategy_detail(tracking)
    summary = build_strategy_summary(detail)
    model_summary = build_model_summary(detail)
    save_csv(detail, DATA_DIR / 'latest_strategy_validation_detail.csv')
    save_csv(summary, DATA_DIR / 'latest_strategy_validation_summary.csv')
    save_csv(model_summary, DATA_DIR / 'latest_recommendation_model_summary.csv')
    save_csv(summary, DATA_DIR / 'latest_strategy_test.csv')
    save_csv(model_summary, DATA_DIR / 'latest_model_test.csv')
    update_workbook(xlsx, tracking, detail, summary, model_summary)
    render_strategy_html(tracking, detail, summary, model_summary)
    update_mobile_links()
    formulas = """v11.4 Google Sheets IMPORTDATA formulas

추천전략검증:
=IMPORTDATA("https://boxinmycat.github.io/stock-report/data/latest_strategy_validation_summary.csv")

추천모델검증:
=IMPORTDATA("https://boxinmycat.github.io/stock-report/data/latest_recommendation_model_summary.csv")

추천추적:
=IMPORTDATA("https://boxinmycat.github.io/stock-report/data/latest_recommendation_tracking.csv")

전략검증상세:
=IMPORTDATA("https://boxinmycat.github.io/stock-report/data/latest_strategy_validation_detail.csv")
"""
    (DOCS / 'GOOGLE_SHEETS_V11_4_FORMULAS.txt').write_text(formulas, encoding='utf-8')
    print('✅ v11.4 추천목록 자동검증 완료')


if __name__ == '__main__':
    main()
